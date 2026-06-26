# Copyright: (c) OpenSpug Organization. https://github.com/openspug/spug
# Copyright: (c) <spug.dev@gmail.com>
# Released under the AGPL-3.0 License.
"""
公共导出工具

提供 Excel 导出的统一能力：工作簿创建、表头样式、列宽、数据行写入、
中文文件名响应头、空数据/超量导出异常处理。

设计目标：
- 统一导出机制，不统一业务字段。每个模块维护自己的 columns 配置。
- 后端导出基于当前筛选条件下的全部数据，而非当前页。
"""
import logging
from datetime import datetime, date
from urllib.parse import quote

from django.http import HttpResponse, JsonResponse

logger = logging.getLogger(__name__)

# 默认导出上限，超过时后端拒绝导出，提示用户缩小筛选范围
DEFAULT_EXPORT_LIMIT = 10000


def build_excel_response(filename, sheet_name, columns, rows):
    """
    构建 Excel 导出响应。

    Args:
        filename: 导出文件名（含扩展名，如 "故障处置记录_xxx.xlsx"）
        sheet_name: 工作表名称
        columns: 列定义，支持两种形式：
            - [(field, title), ...] 元组列表
            - [{'key': field, 'title': title}, ...] 字典列表
        rows: dict 列表，每个 dict 是一条记录，key 与 columns 的 field 对应

    Returns:
        HttpResponse: Content-Type 为 Excel 二进制流
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 归一化 columns 为 [(key, title), ...]
    normalized_columns = _normalize_columns(columns)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 表头样式
    header_font = Font(name='宋体', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    body_font = Font(name='宋体', size=10)
    body_align = Alignment(vertical='center', wrap_text=True)

    # 写表头
    for col_idx, (_key, title) in enumerate(normalized_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # 写数据行
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, (key, _title) in enumerate(normalized_columns, start=1):
            value = _get_cell_value(row_data, key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = border

    # 自适应列宽（基于表头与内容长度估算，设上下限）
    for col_idx, (key, title) in enumerate(normalized_columns, start=1):
        max_len = len(str(title)) * 2  # 中文按 2 计
        # 抽样前 200 行估算内容宽度，避免大数据下遍历过慢
        for row_data in rows[:200]:
            val = _get_cell_value(row_data, key)
            if val is not None and val != '':
                max_len = max(max_len, min(len(str(val)) * 2, 60))
        width = max(max_len + 4, 10)
        width = min(width, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结表头
    ws.freeze_panes = 'A2'

    # 写入内存
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # 中文文件名使用 RFC 5987 编码，兼容主流浏览器
    response['Content-Disposition'] = "attachment; filename*=UTF-8''%s" % quote(filename)
    return response


def build_export_error_response(message, status=400):
    """
    构建导出错误响应（JSON）。

    前端 http 拦截器已处理二进制响应中的 JSON 错误，
    返回 JsonResponse 即可被前端正确解析为错误提示。
    """
    return JsonResponse({'error': message}, status=status)


def check_export_limit(queryset, limit=DEFAULT_EXPORT_LIMIT):
    """
    检查导出数量是否超过上限。

    Args:
        queryset: Django QuerySet
        limit: 上限，默认 10000

    Returns:
        (count, error_response_or_none)
        若超过上限，返回 (count, JsonResponse)；否则返回 (count, None)。
    """
    count = queryset.count()
    if count > limit:
        msg = '导出数据超过 %d 条，请缩小筛选范围后重试' % limit
        return count, build_export_error_response(msg)
    return count, None


def _normalize_columns(columns):
    """将 columns 归一化为 [(key, title), ...] 形式"""
    normalized = []
    for col in columns:
        if isinstance(col, (list, tuple)):
            normalized.append((col[0], col[1]))
        elif isinstance(col, dict):
            normalized.append((col['key'], col['title']))
        else:
            raise ValueError('columns 元素必须是 (field, title) 元组或 {"key","title"} 字典')
    return normalized


def _get_cell_value(row_data, key):
    """从行数据中取值，处理 model 实例、dict、None、日期时间等"""
    if row_data is None:
        return ''
    # 支持 dict
    if isinstance(row_data, dict):
        value = row_data.get(key, '')
    else:
        # 支持 model 实例 / 对象属性
        value = getattr(row_data, key, '')

    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    return value

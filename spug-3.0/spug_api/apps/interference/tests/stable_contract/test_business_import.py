# -*- coding: utf-8 -*-
"""干扰管理双业务类型 Excel 导入 stable_contract 测试。

覆盖：模板下载 / 表头乱序与标题行 / 合并单元格拒绝 / 斜线样式与组合表头 /
缺列重复列 / 日期与持续时间与高度单位校验 / 公式拒绝 / 未识别列警告 /
空行截断 / 文件内与数据库重复 / 预校验不写库 / 确认导入创建与整体回滚 /
租户隔离 / 权限拒绝 / 幂等 / 行数上限 / 审计日志 / 错误报告。
"""
import io
import json
from datetime import datetime
from unittest import mock

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side

from apps.utils.test_helpers import make_user, make_client, setup_test_env
from apps.interference.models import (
    AirInterferenceRecord,
    BridgeInterferenceRecord,
    format_decimal,
)
from apps.logs.models import AuditLog

XLSX_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

BRIDGE_HEADER = ['日期时间', '航班号', '机号', '机型', '位置/机位', '频率', '现象',
                 '处置方式', '原因分析', '备注']
AIR_HEADER = ['日期时间', '航班号', '机型', '航线', '被扰频率',
              '告警高度', '高度单位', '告警航段', '持续时间', '现象', '处置方式', '原因分析']


def build_excel(header, data_rows, header_row=1, pre_rows=(), merges=(),
                formulas=None, border_cells=()):
    """构造测试用 Excel。

    header: list[str] 表头（从第 1 列开始）
    data_rows: list[list] 数据行（与表头同列数，可含 datetime 对象）
    header_row: 表头所在行号（1 起始）
    pre_rows: 表头前的标题/说明行
    merges: 合并区域列表，如 ['A6:D6']，行号基于最终工作表
    formulas: dict (row, col) -> 公式文本（绝对坐标）
    border_cells: list (row, col) 设置斜线边框（仅样式）
    """
    wb = Workbook()
    ws = wb.active
    row_cursor = 1
    for text in pre_rows:
        ws.cell(row=row_cursor, column=1, value=text)
        row_cursor += 1
    while row_cursor < header_row:
        row_cursor += 1
    for col_idx, label in enumerate(header, start=1):
        ws.cell(row=header_row, column=col_idx, value=label)
    for offset, values in enumerate(data_rows, start=1):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=header_row + offset, column=col_idx, value=value)
    for coord in merges:
        ws.merge_cells(coord)
    for (row_idx, col_idx), formula in (formulas or {}).items():
        ws.cell(row=row_idx, column=col_idx, value=formula)
    for row_idx, col_idx in border_cells:
        ws.cell(row=row_idx, column=col_idx).border = Border(
            diagonal=Side(border_style='thin'), diagonalDown=True, diagonalUp=True)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def upload_xlsx(client, url, data, filename='导入.xlsx', extra=None):
    file = SimpleUploadedFile(filename, data, content_type=XLSX_CONTENT_TYPE)
    payload = {'file': file}
    if extra:
        payload.update(extra)
    return client.post(url, payload)


BRIDGE_KEY_TO_LABEL = dict(zip(
    ['datetime', 'flight_number', 'aircraft_no', 'aircraft_type', 'location',
     'frequency', 'phenomenon', 'handling_method', 'cause_analysis', 'remark'],
    BRIDGE_HEADER))
AIR_KEY_TO_LABEL = dict(zip(
    ['datetime', 'flight_number', 'aircraft_type', 'route',
     'alert_form', 'alert_altitude', 'alert_altitude_unit',
     'alert_segment', 'duration', 'phenomenon', 'handling_method', 'cause_analysis'],
    AIR_HEADER))


def _normalize_overrides(overrides, key_to_label):
    """允许用模型字段名或中文表头名覆盖单元格值。"""
    return {key_to_label.get(k, k): v for k, v in overrides.items()}


def bridge_row(**overrides):
    values = ['2026-08-01 10:00', 'CA1234', 'B-2026', 'A320', '007',
              '118.6', '甚高频出现杂音', '通知机务排查', '地面电源车干扰', '备注内容']
    data = dict(zip(BRIDGE_HEADER, values))
    data.update(_normalize_overrides(overrides, BRIDGE_KEY_TO_LABEL))
    return [data[key] for key in BRIDGE_HEADER]


def air_row(**overrides):
    values = ['2026-08-02 09:30', 'MU5678', 'B738', 'KMG-SHA',
              '高度告警', '1200', '米', '五边进近', '01:30:00', '低高度告警', '', '']
    data = dict(zip(AIR_HEADER, values))
    data.update(_normalize_overrides(overrides, AIR_KEY_TO_LABEL))
    return [data[key] for key in AIR_HEADER]


class ImportTemplateTest(TestCase):
    """两类模板下载：可被机器稳定识别表头，文本列格式正确。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('imp_tpl', ['interference.interference.view'])
        self.client = make_client(self.user)

    def _check_template(self, url, header, title):
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # 前 20 行内存在一行完整表头
        header_row = None
        for row_idx in range(1, 21):
            cells = [str(c.value) if c.value else '' for c in ws[row_idx]]
            if all(name in cells for name in header):
                header_row = row_idx
                break
        self.assertIsNotNone(header_row, '模板中找不到可识别的表头行')
        # 表头列设置为文本格式
        for col_idx in range(1, len(header) + 1):
            self.assertEqual(ws.cell(row=header_row + 1, column=col_idx).number_format, '@')
        # 标题与说明存在且位于表头之前
        self.assertTrue(any(title in str(c.value) for c in ws[1]))
        return ws

    def test_bridge_template(self):
        self._check_template('/interference/bridge/import/template/', BRIDGE_HEADER,
                             '地面无线电通信异常/干扰')

    def test_air_template(self):
        self._check_template('/interference/air/import/template/', AIR_HEADER, '空中干扰')

    def test_template_download_audited(self):
        self.client.get('/interference/bridge/import/template/')
        self.assertTrue(AuditLog.objects.filter(
            action='export', detail__contains='import_template').exists())

    def test_template_requires_view_permission(self):
        no_perm = make_user('imp_tpl_nopm', [])
        resp = make_client(no_perm).get('/interference/bridge/import/template/')
        self.assertTrue(resp.json().get('error'))


class BridgeImportValidateTest(TestCase):
    """地面预校验：标题行+乱序表头可识别；预校验不写业务数据。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('imp_val', ['interference.interference.view',
                                          'interference.interference.add'])
        self.client = make_client(self.user)

    def _validate(self, data, filename='导入.xlsx'):
        return upload_xlsx(self.client, '/interference/bridge/import/validate/',
                           data, filename)

    def test_validate_ok_with_title_and_shuffled_header(self):
        # 表头乱序（逆序排列）+ 表头前有标题/说明行，仍可识别导入
        shuffled_header = BRIDGE_HEADER[::-1]
        row = bridge_row()
        shuffled_row = row[::-1]
        data = build_excel(shuffled_header, [shuffled_row], header_row=5,
                           pre_rows=['某某机场运行保障部', '廊桥干扰登记表', '填表人：张三'])
        resp = self._validate(data)
        body = resp.json()['data']
        self.assertEqual(body['total_rows'], 1)
        self.assertEqual(body['valid_count'], 1)
        self.assertEqual(body['error_count'], 0, body['errors'])
        self.assertTrue(body['validate_token'])
        # 预校验不产生业务数据
        self.assertFalse(BridgeInterferenceRecord.objects.exists())

    def test_validate_writes_nothing_with_multiple_rows(self):
        data = build_excel(BRIDGE_HEADER, [bridge_row(), bridge_row(flight_number='CA9999')])
        resp = self._validate(data)
        self.assertEqual(resp.json()['data']['total_rows'], 2)
        self.assertFalse(BridgeInterferenceRecord.objects.exists())

    def test_xls_rejected(self):
        data = build_excel(BRIDGE_HEADER, [bridge_row()])
        resp = self._validate(data, filename='导入.xls')
        self.assertIn('.xlsx', resp.json()['error'])

    def test_rows_limit_fail_closed(self):
        data = build_excel(BRIDGE_HEADER, [bridge_row(flight_number=f'CA{i:04d}')
                                           for i in range(3)])
        with override_settings(INTERFERENCE_IMPORT_MAX_ROWS=2):
            resp = self._validate(data)
        self.assertIn('上限', resp.json()['error'])


class BridgeImportCommitTest(TestCase):
    """地面确认导入：真实建记录、字段/租户/审计正确。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('imp_cmt', ['interference.interference.view',
                                          'interference.interference.add'])
        self.client = make_client(self.user)
        self.user.tenant_id = 'tenant_import'
        self.user.save()

    def _validate(self, data):
        return upload_xlsx(self.client, '/interference/bridge/import/validate/', data)

    def _commit(self, data, token):
        return upload_xlsx(self.client, '/interference/bridge/import/commit/', data,
                           extra={'validate_token': token})

    def test_commit_creates_records_with_correct_fields(self):
        data = build_excel(BRIDGE_HEADER, [bridge_row()])
        body = self._validate(data).json()['data']
        resp = self._commit(data, body['validate_token'])
        self.assertFalse(resp.json().get('error'), resp.json())
        self.assertEqual(resp.json()['data']['imported_count'], 1)

        record = BridgeInterferenceRecord.objects.get()
        self.assertEqual(record.flight_number, 'CA1234')
        # 前导零按文本保留（位置/机位列为文本格式）
        self.assertEqual(record.location, '007')
        self.assertEqual(record.aircraft_no, 'B-2026')
        self.assertEqual(record.frequency, '118.6')
        self.assertEqual(record.phenomenon, '甚高频出现杂音')
        self.assertEqual(record.remark, '备注内容')
        self.assertEqual(record.created_by, self.user)
        self.assertEqual(record.tenant_id, 'tenant_import')

    def test_excel_date_cell_accepted(self):
        row = bridge_row()
        row[0] = datetime(2026, 8, 1, 10, 0, 0)
        data = build_excel(BRIDGE_HEADER, [row])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])
        resp = self._commit(data, body['validate_token'])
        self.assertFalse(resp.json().get('error'), resp.json())
        record = BridgeInterferenceRecord.objects.get()
        self.assertEqual(record.datetime.strftime('%Y-%m-%d %H:%M:%S'),
                         '2026-08-01 10:00:00')

    def test_date_only_text_normalized(self):
        row = bridge_row(**{'日期时间': '2026-08-01'})
        data = build_excel(BRIDGE_HEADER, [row])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])
        self._commit(data, body['validate_token'])
        record = BridgeInterferenceRecord.objects.get()
        self.assertEqual(record.datetime.strftime('%Y-%m-%d %H:%M:%S'),
                         '2026-08-01 00:00:00')

    def test_date_only_variants_accepted(self):
        """仅日期与点分隔日期+时分兼容：横杠/斜杠/点/中文写法与 Excel 日期单元格。"""
        variants = [
            ('2026-08-01', 'CA8001', '2026-08-01 00:00'),
            ('2026/8/1', 'CA8002', '2026-08-01 00:00'),
            ('2026.8.1', 'CA8003', '2026-08-01 00:00'),
            ('2026年8月1日', 'CA8004', '2026-08-01 00:00'),
            # 点分隔日期 + 时分
            ('2024.7.8 17:15', 'CA8005', '2024-07-08 17:15'),
            ('2026/8/1 8:05', 'CA8006', '2026-08-01 08:05'),
            ('2026-08-03 09:15:59', 'CA8007', '2026-08-03 09:15'),
        ]
        for text, flight, expected in variants:
            row = bridge_row(**{'日期时间': text, '航班号': flight})
            data = build_excel(BRIDGE_HEADER, [row])
            body = self._validate(data).json()['data']
            self.assertEqual(body['error_count'], 0,
                             f'{text} 应可导入: {body["errors"]}')
            self._commit(data, body['validate_token'])
            record = BridgeInterferenceRecord.objects.get(flight_number=flight)
            self.assertEqual(record.datetime.strftime('%Y-%m-%d %H:%M'), expected)

        # Excel 日期单元格（仅日期，无时分）
        row = bridge_row()
        row[0] = datetime(2026, 8, 2).date()
        data = build_excel(BRIDGE_HEADER, [row])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])
        self._commit(data, body['validate_token'])
        record = BridgeInterferenceRecord.objects.get(flight_number='CA1234')
        self.assertEqual(record.datetime.strftime('%Y-%m-%d %H:%M'), '2026-08-02 00:00')

    def test_empty_flight_no_and_type_importable(self):
        """航班号/机号/机型允许为空（仅日期时间与现象必填）。"""
        row = bridge_row(**{'航班号': '', '机号': '', '机型': ''})
        data = build_excel(BRIDGE_HEADER, [row])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])
        self._commit(data, body['validate_token'])
        record = BridgeInterferenceRecord.objects.get()
        self.assertEqual(record.flight_number, '')
        self.assertEqual(record.aircraft_no, '')
        self.assertEqual(record.aircraft_type, '')

    def test_minute_precision_accepted_and_seconds_truncated(self):
        # 日期时间到分钟即可；精确到秒的输入被接受但截断到分钟
        row = bridge_row(**{'日期时间': '2026-08-01 10:30'})
        data = build_excel(BRIDGE_HEADER, [row])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])
        self._commit(data, body['validate_token'])
        record = BridgeInterferenceRecord.objects.get(flight_number='CA1234')
        self.assertEqual(record.datetime.strftime('%Y-%m-%d %H:%M'), '2026-08-01 10:30')

        row = bridge_row(**{'日期时间': '2026-08-02 09:15:59',
                            '航班号': 'CA2222'})
        data = build_excel(BRIDGE_HEADER, [row])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])
        self._commit(data, body['validate_token'])
        record = BridgeInterferenceRecord.objects.get(flight_number='CA2222')
        self.assertEqual(record.datetime.strftime('%Y-%m-%d %H:%M'), '2026-08-02 09:15')

    def test_commit_any_row_failure_rolls_back_all(self):
        data = build_excel(BRIDGE_HEADER, [bridge_row(flight_number='CA0001'),
                                           bridge_row(flight_number='CA0002')])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0)

        original_create = BridgeInterferenceRecord.objects.create

        def fake_create(**kwargs):
            if kwargs.get('flight_number') == 'CA0002':
                raise RuntimeError('模拟数据库写入失败')
            return original_create(**kwargs)

        with mock.patch.object(BridgeInterferenceRecord.objects, 'create',
                               side_effect=fake_create):
            resp = self._commit(data, body['validate_token'])
        self.assertTrue(resp.json().get('error'))
        self.assertIn('回滚', resp.json()['error'])
        self.assertEqual(BridgeInterferenceRecord.objects.count(), 0)


class AirImportFlowTest(TestCase):
    """空中导入：数值字段规则、可选字段、字段分离。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('imp_air', ['interference.interference.view',
                                          'interference.interference.add'])
        self.client = make_client(self.user)

    def _validate(self, data):
        return upload_xlsx(self.client, '/interference/air/import/validate/', data)

    def _commit(self, data, token):
        return upload_xlsx(self.client, '/interference/air/import/commit/', data,
                           extra={'validate_token': token})

    def test_full_flow_with_numeric_fields(self):
        data = build_excel(AIR_HEADER, [air_row()])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])
        resp = self._commit(data, body['validate_token'])
        self.assertFalse(resp.json().get('error'), resp.json())
        record = AirInterferenceRecord.objects.get()
        # 高度与单位分别保存，不换算
        self.assertEqual(format_decimal(record.alert_altitude), '1200')
        self.assertEqual(record.alert_altitude_unit, 'm')
        # 持续时间 HH:MM:SS -> 秒（无损），单位 s
        self.assertEqual(format_decimal(record.duration), '5400')
        self.assertEqual(record.duration_unit, 's')
        # 现象 / 处置方式分别入库
        self.assertEqual(record.phenomenon, '低高度告警')
        self.assertEqual(record.handling_method, '')
        self.assertEqual(record.cause_analysis, '')

    def test_altitude_unit_alias_and_no_conversion(self):
        row = air_row(**{'告警高度': '1200', '高度单位': '英尺'})
        data = build_excel(AIR_HEADER, [row])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])
        self._commit(data, body['validate_token'])
        record = AirInterferenceRecord.objects.get()
        self.assertEqual(record.alert_altitude_unit, 'ft')
        self.assertEqual(format_decimal(record.alert_altitude), '1200')

    def test_first_registration_allows_empty_handling_fields(self):
        row = air_row(**{'处置方式': '', '原因分析': ''})
        data = build_excel(AIR_HEADER, [row])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])
        resp = self._commit(data, body['validate_token'])
        self.assertFalse(resp.json().get('error'), resp.json())
        record = AirInterferenceRecord.objects.get()
        self.assertEqual(record.handling_method, '')

    def test_cross_business_file_rejected(self):
        # 空中文件上传到地面入口：必须拒绝，不允许静默丢列
        data = build_excel(AIR_HEADER, [air_row()])
        resp = upload_xlsx(self.client, '/interference/bridge/import/validate/', data)
        error = resp.json().get('error')
        self.assertTrue(error)
        self.assertIn('空中干扰', error)
        self.assertFalse(BridgeInterferenceRecord.objects.exists())

        # 地面文件上传到空中入口：同样拒绝
        bridge_data = build_excel(BRIDGE_HEADER, [bridge_row()])
        resp = upload_xlsx(self.client, '/interference/air/import/validate/', bridge_data)
        error = resp.json().get('error')
        self.assertTrue(error)
        self.assertIn('地面无线电通信异常/干扰', error)


class ImportHeaderGuardTest(TestCase):
    """表头识别规则：合并/斜线/缺列/重复列/未知列/空行截断。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('imp_hdr', ['interference.interference.view',
                                          'interference.interference.add'])
        self.client = make_client(self.user)

    def _validate(self, data):
        return upload_xlsx(self.client, '/interference/bridge/import/validate/', data)

    def test_merged_header_cell_rejected(self):
        # 表头行第 6 行：A6:C6 合并单元格承载表头
        data = build_excel(BRIDGE_HEADER, [bridge_row()], header_row=6,
                           pre_rows=['标题'] * 4, merges=['A6:C6'])
        resp = self._validate(data)
        error = resp.json().get('error')
        self.assertTrue(error)
        self.assertIn('合并单元格', error)
        self.assertIn('6', error)

    def test_merged_data_row_rejected_with_row_number(self):
        data = build_excel(BRIDGE_HEADER, [bridge_row()], header_row=1,
                           merges=['A2:C2'])
        resp = self._validate(data)
        body = resp.json()['data']
        self.assertEqual(body['error_count'], 1)
        error = body['errors'][0]
        self.assertEqual(error['row'], 2)
        self.assertIn('合并单元格', error['message'])

    def test_slash_combined_header_rejected(self):
        # 一个单元格用斜线表达两个字段（航班号/机号）：必须提示拆分为两列
        header = ['日期时间', '航班号/机号', '机型', '位置/机位', '频率', '现象', '备注']
        row = ['2026-08-01 10:00:00', 'CA1234 / B-2026', 'A320', 'T2廊桥007号机位',
               '118.6', '甚高频出现杂音', '']
        data = build_excel(header, [row])
        resp = self._validate(data)
        error = resp.json().get('error')
        self.assertTrue(error)
        self.assertIn('拆分', error)

    def test_slash_border_header_still_recognized(self):
        # 斜线边框仅是样式，不应影响普通表头识别
        data = build_excel(BRIDGE_HEADER, [bridge_row()], border_cells=[(1, 1), (1, 2)])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])
        self.assertEqual(body['valid_count'], 1)

    def test_missing_required_header_rejected(self):
        header = [name for name in BRIDGE_HEADER if name != '现象']
        row = [v for k, v in zip(BRIDGE_HEADER, bridge_row()) if k != '现象']
        data = build_excel(header, [row])
        resp = self._validate(data)
        error = resp.json().get('error')
        self.assertTrue(error)
        self.assertIn('现象', error)
        self.assertIn('缺少必需表头', error)

    def test_duplicate_header_rejected(self):
        header = BRIDGE_HEADER + ['日期时间']
        row = bridge_row() + ['2026-08-02 10:00:00']
        data = build_excel(header, [row])
        resp = self._validate(data)
        error = resp.json().get('error')
        self.assertTrue(error)
        self.assertIn('重复表头', error)

    def test_unknown_column_warned_but_row_valid(self):
        header = BRIDGE_HEADER + ['填表人']
        row = bridge_row() + ['张三']
        data = build_excel(header, [row])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])
        self.assertEqual(body['warning_count'], 1)
        self.assertIn('填表人', body['warnings'][0]['message'])
        self.assertIn('不会导入', body['warnings'][0]['message'])

    def test_legacy_location_only_column_mapped_to_merged_field(self):
        """旧格式文件仅有「位置」列（或「机位」列）时仍可映射到合并字段。"""
        header = ['日期时间', '航班号', '机号', '机型', '位置', '频率', '现象', '备注']
        row = ['2026-08-01 10:00:00', 'CA1234', 'B-2026', 'A320', 'T2廊桥',
               '118.6', '甚高频出现杂音', '']
        data = build_excel(header, [row])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])
        self.assertEqual(body['valid_count'], 1)

    def test_legacy_both_location_and_stand_no_rejected_as_duplicate(self):
        """旧格式文件同时含「位置」「机位」两列：同字段重复表头，拒绝导入。"""
        header = ['日期时间', '航班号', '机号', '机型', '位置', '机位', '频率', '现象', '备注']
        row = ['2026-08-01 10:00:00', 'CA1234', 'B-2026', 'A320', 'T2廊桥',
               '007号机位', '118.6', '甚高频出现杂音', '']
        data = build_excel(header, [row])
        error = self._validate(data).json().get('error')
        self.assertTrue(error)
        self.assertIn('重复表头', error)
        self.assertIn('位置/机位', error)

    def test_blank_row_stops_reading(self):
        data = build_excel(BRIDGE_HEADER, [bridge_row()], header_row=1)
        # 数据行后为空行（第 3 行），其后为页脚说明（第 4 行）
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        ws.cell(row=4, column=1, value='页脚：以上内容无需填写')
        buf = io.BytesIO()
        wb.save(buf)
        body = self._validate(buf.getvalue()).json()['data']
        self.assertEqual(body['total_rows'], 1)

    def test_header_beyond_20_rows_rejected(self):
        data = build_excel(BRIDGE_HEADER, [bridge_row()], header_row=25,
                           pre_rows=[f'说明{i}' for i in range(24)])
        resp = self._validate(data)
        self.assertIn('表头', resp.json().get('error'))


class ImportRowValidationTest(TestCase):
    """行级校验：日期/持续时间/高度单位/公式/必填/长度，与手工表单同一套规则。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('imp_row', ['interference.interference.view',
                                          'interference.interference.add'])
        self.client = make_client(self.user)
        self.air_client_url = '/interference/air/import/validate/'

    def _validate_bridge(self, data):
        return upload_xlsx(self.client, '/interference/bridge/import/validate/', data)

    def _validate_air(self, data):
        return upload_xlsx(self.client, self.air_client_url, data)

    def _errors(self, resp):
        return resp.json()['data']['errors']

    def test_bad_datetime_text_rejected(self):
        # 斜杠日期+时分已是合法写法；此处验证真正无法识别的文本被拒绝
        for bad in ('2026/08/01 上午', '08/01/2026', 'not-a-date'):
            row = bridge_row(**{'日期时间': bad})
            errors = self._errors(self._validate_bridge(build_excel(BRIDGE_HEADER, [row])))
            self.assertTrue(any(e['field'] == '日期时间' for e in errors),
                            f'{bad} 应被拒绝: {errors}')

    def test_bad_duration_rejected_with_hint(self):
        row = air_row(**{'持续时间': '90分钟'})
        errors = self._errors(self._validate_air(build_excel(AIR_HEADER, [row])))
        self.assertTrue(any('HH:MM:SS' in e['message'] for e in errors), errors)

        row = air_row(**{'持续时间': '1:30'})
        errors = self._errors(self._validate_air(build_excel(AIR_HEADER, [row])))
        self.assertTrue(any('HH:MM:SS' in e['message'] for e in errors), errors)

    def test_zero_duration_rejected(self):
        row = air_row(**{'持续时间': '00:00:00'})
        errors = self._errors(self._validate_air(build_excel(AIR_HEADER, [row])))
        self.assertTrue(any('大于0' in e['message'] for e in errors), errors)

    def test_bad_altitude_unit_rejected(self):
        row = air_row(**{'高度单位': '公里'})
        errors = self._errors(self._validate_air(build_excel(AIR_HEADER, [row])))
        self.assertTrue(any(e['field'] == '高度单位' for e in errors), errors)

    def test_altitude_without_unit_rejected(self):
        row = air_row(**{'高度单位': ''})
        errors = self._errors(self._validate_air(build_excel(AIR_HEADER, [row])))
        self.assertTrue(any('高度单位' in e['message'] for e in errors), errors)

    def test_unit_without_altitude_warns(self):
        row = air_row(**{'告警高度': ''})
        resp = self._validate_air(build_excel(AIR_HEADER, [row])).json()['data']
        self.assertEqual(resp['error_count'], 0)
        self.assertEqual(resp['warning_count'], 1)

    def test_negative_altitude_rejected(self):
        row = air_row(**{'告警高度': '-100'})
        errors = self._errors(self._validate_air(build_excel(AIR_HEADER, [row])))
        self.assertTrue(any('告警高度' in e['message'] for e in errors), errors)

    def test_formula_cell_rejected(self):
        data = build_excel(BRIDGE_HEADER, [bridge_row()], formulas={(2, 7): '=SUM(1+1)'})
        errors = self._errors(self._validate_bridge(data))
        self.assertTrue(any('公式' in e['message'] for e in errors), errors)
        self.assertEqual(errors[0]['row'], 2)

    def test_missing_required_cell_rejected(self):
        row = bridge_row(**{'现象': ''})
        errors = self._errors(self._validate_bridge(build_excel(BRIDGE_HEADER, [row])))
        self.assertTrue(any(e['message'] == '请输入现象' for e in errors), errors)

    def test_over_length_text_rejected(self):
        row = bridge_row(**{'航班号': 'X' * 150})
        errors = self._errors(self._validate_bridge(build_excel(BRIDGE_HEADER, [row])))
        self.assertTrue(any('长度不能超过100' in e['message'] for e in errors), errors)


class ImportDuplicateTest(TestCase):
    """文件内重复与数据库重复均拒绝；重复键按业务定义且租户隔离。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('imp_dup', ['interference.interference.view',
                                          'interference.interference.add'])
        self.client = make_client(self.user)

    def _validate(self, data):
        return upload_xlsx(self.client, '/interference/bridge/import/validate/', data)

    def test_in_file_duplicate_rejected(self):
        data = build_excel(BRIDGE_HEADER, [bridge_row(), bridge_row()])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 1)
        error = body['errors'][0]
        self.assertIn('文件内第', error['message'])
        self.assertIn('机位', error['message'])
        self.assertIn('频率', error['message'])

    def test_in_file_different_frequency_not_duplicate(self):
        # 地面重复键：日期时间+航班号+机位+频率；仅频率不同不算重复
        data = build_excel(BRIDGE_HEADER, [bridge_row(), bridge_row(frequency='121.5')])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])

    def test_db_duplicate_rejected(self):
        BridgeInterferenceRecord.objects.create(
            tenant_id=self.user.tenant_id,
            datetime='2026-08-01 10:00:00', flight_number='CA1234',
            location='007', frequency='118.6', phenomenon='已有记录',
            created_by=self.user)
        data = build_excel(BRIDGE_HEADER, [bridge_row()])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 1)
        self.assertIn('数据库中已存在', body['errors'][0]['message'])
        self.assertFalse(BridgeInterferenceRecord.objects.filter(
            phenomenon='甚高频出现杂音').exists())

    def test_db_duplicate_check_tenant_isolated(self):
        BridgeInterferenceRecord.objects.create(
            tenant_id='tenant_other',
            datetime='2026-08-01 10:00:00', flight_number='CA1234',
            location='007', frequency='118.6', phenomenon='别家记录',
            created_by=self.user)
        data = build_excel(BRIDGE_HEADER, [bridge_row()])
        body = self._validate(data).json()['data']
        self.assertEqual(body['error_count'], 0, body['errors'])


class ImportPermissionTenantTest(TestCase):
    """权限与租户：写入接口需 add 权限；数据按登录用户租户写入。"""

    def setUp(self):
        setup_test_env(self)
        self.viewer = make_user('imp_view', ['interference.interference.view'])
        self.adder = make_user('imp_add', ['interference.interference.view',
                                           'interference.interference.add'])
        self.adder.tenant_id = 'tenant_a'
        self.adder.save()

    def test_validate_and_commit_require_add_permission(self):
        client = make_client(self.viewer)
        data = build_excel(BRIDGE_HEADER, [bridge_row()])
        resp = upload_xlsx(client, '/interference/bridge/import/validate/', data)
        self.assertTrue(resp.json().get('error'))
        resp = upload_xlsx(client, '/interference/bridge/import/commit/', data,
                           extra={'validate_token': 'x'})
        self.assertTrue(resp.json().get('error'))
        self.assertFalse(BridgeInterferenceRecord.objects.exists())

    def test_unauthenticated_rejected(self):
        from django.test import Client
        data = build_excel(BRIDGE_HEADER, [bridge_row()])
        resp = Client().post('/interference/bridge/import/validate/', {'file': data})
        self.assertTrue(resp.json().get('error'))

    def test_commit_tenant_follows_login_user(self):
        client = make_client(self.adder)
        data = build_excel(BRIDGE_HEADER, [bridge_row()])
        body = upload_xlsx(client, '/interference/bridge/import/validate/', data).json()['data']
        resp = upload_xlsx(client, '/interference/bridge/import/commit/', data,
                           extra={'validate_token': body['validate_token']})
        self.assertFalse(resp.json().get('error'), resp.json())
        record = BridgeInterferenceRecord.objects.get()
        self.assertEqual(record.tenant_id, 'tenant_a')
        # 其他租户用户不可见（TenantModelManager 自动过滤）
        other = make_user('imp_other', ['interference.interference.view'])
        other.tenant_id = 'tenant_b'
        other.save()
        resp = make_client(other).get('/interference/bridge/')
        self.assertEqual(resp.json()['data']['total'], 0)


class ImportIdempotencyTest(TestCase):
    """幂等：凭证一次性、绑定用户/业务/文件内容；重复提交被拒绝。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('imp_idem', ['interference.interference.view',
                                           'interference.interference.add'])
        self.client = make_client(self.user)
        self.data = build_excel(BRIDGE_HEADER, [bridge_row()])

    def _validate(self, data=None):
        return upload_xlsx(self.client, '/interference/bridge/import/validate/',
                           data or self.data).json()['data']

    def _commit(self, data, token):
        return upload_xlsx(self.client, '/interference/bridge/import/commit/', data,
                           extra={'validate_token': token})

    def test_commit_without_token_rejected(self):
        resp = self._commit(self.data, '')
        self.assertIn('预校验凭证', resp.json()['error'])
        self.assertFalse(BridgeInterferenceRecord.objects.exists())

    def test_double_commit_rejected(self):
        token = self._validate()['validate_token']
        resp = self._commit(self.data, token)
        self.assertFalse(resp.json().get('error'), resp.json())
        self.assertEqual(BridgeInterferenceRecord.objects.count(), 1)
        # 第二次提交（双击/网络重试）被拒绝，且不产生重复数据
        token2 = self._validate()['validate_token']
        resp = self._commit(self.data, token2)
        self.assertTrue(resp.json().get('error'))
        self.assertEqual(BridgeInterferenceRecord.objects.count(), 1)

    def test_commit_with_mismatched_file_rejected(self):
        token = self._validate()['validate_token']
        other_data = build_excel(BRIDGE_HEADER, [bridge_row(flight_number='CA8888')])
        resp = self._commit(other_data, token)
        self.assertIn('不一致', resp.json()['error'])
        self.assertFalse(BridgeInterferenceRecord.objects.exists())


class ImportAuditTest(TestCase):
    """导入审计：记录导入人/类型/文件名/SHA-256/行数/成功失败数/创建ID，无单元格内容。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('imp_audit', ['interference.interference.view',
                                            'interference.interference.add'])
        self.client = make_client(self.user)

    def test_commit_success_audit(self):
        data = build_excel(BRIDGE_HEADER, [bridge_row()])
        body = upload_xlsx(self.client, '/interference/bridge/import/validate/',
                           data).json()['data']
        upload_xlsx(self.client, '/interference/bridge/import/commit/', data,
                    extra={'validate_token': body['validate_token']})
        log = AuditLog.objects.filter(action='import').latest('id')
        self.assertEqual(log.username, self.user.username)
        detail = json.loads(log.detail)
        self.assertEqual(detail['record_type'], 'bridge_interference')
        self.assertEqual(detail['file_name'], '导入.xlsx')
        self.assertEqual(len(detail['file_sha256']), 64)
        self.assertEqual(detail['total_rows'], 1)
        self.assertEqual(detail['success_count'], 1)
        self.assertEqual(detail['fail_count'], 0)
        self.assertEqual(len(detail['created_ids']), 1)
        # 不得将完整 Excel 内容写入审计
        self.assertNotIn('甚高频出现杂音', log.detail)
        self.assertNotIn('CA1234', log.detail)

    def test_commit_failure_audit(self):
        row = bridge_row(**{'现象': ''})
        data = build_excel(BRIDGE_HEADER, [row])
        body = upload_xlsx(self.client, '/interference/bridge/import/validate/',
                           data).json()['data']
        upload_xlsx(self.client, '/interference/bridge/import/commit/', data,
                    extra={'validate_token': body['validate_token']})
        log = AuditLog.objects.filter(action='import').latest('id')
        detail = json.loads(log.detail)
        self.assertEqual(detail['success_count'], 0)
        self.assertEqual(detail['fail_count'], 1)
        self.assertFalse(log.is_success)


class ImportErrorReportTest(TestCase):
    """错误报告：原文件列 + Excel行号 + 错误原因；仅包含错误行。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('imp_rpt', ['interference.interference.view',
                                          'interference.interference.add'])
        self.client = make_client(self.user)

    def test_error_report_content(self):
        header = BRIDGE_HEADER + ['填表人']
        rows = [bridge_row(), bridge_row(datetime='bad-date') + ['李四']]
        data = build_excel(header, rows)
        resp = upload_xlsx(self.client, '/interference/bridge/import/error-report/', data)
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers, header + ['Excel行号', '错误原因'])
        # 仅错误行
        self.assertEqual(ws.max_row, 2)
        values = [c.value for c in ws[2]]
        self.assertEqual(values[0], 'bad-date')
        self.assertEqual(values[header.index('填表人')], '李四')
        self.assertIn('日期时间', values[-1])

    def test_error_report_without_errors_rejected(self):
        data = build_excel(BRIDGE_HEADER, [bridge_row()])
        resp = upload_xlsx(self.client, '/interference/bridge/import/error-report/', data)
        self.assertTrue(resp.json().get('error'))

# -*- coding: utf-8 -*-
"""空中干扰记录 stable_contract 测试。

覆盖：创建/字段校验（含告警高度与持续时间的单位和正数校验）/编辑/软删除/
租户隔离/附件归属隔离/导出字段/权限拒绝/状态流转必填规则（处置需处置方式、
关闭需原因分析）/数据库约束兜底。
"""
import json
import tempfile
from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings

from apps.utils.test_helpers import make_user, make_client, setup_test_env
from apps.interference.models import AirInterferenceRecord, BridgeInterferenceRecord
from apps.evidence.models import EvidenceAttachment


def air_data(**overrides):
    data = {
        'datetime': '2026-08-02 14:30:00',
        'flight_number': 'MU5678',
        'aircraft_type': 'B738',
        'route': 'HFE-VVO',
        'alert_form': 'TCAS RA',
        'alert_altitude': 1200,
        'alert_altitude_unit': 'm',
        'alert_segment': '进场下降段',
        'duration': 45,
        'duration_unit': 's',
        'phenomenon': '下降过程中出现TCAS RA告警',
        'handling_method': '',
        'cause_analysis': '',
    }
    data.update(overrides)
    return data


class AirInterferenceAuthTest(TestCase):
    """权限拒绝。"""

    def setUp(self):
        setup_test_env(self)
        self.v = make_user('aiv_view', ['interference.interference.view'])
        self.n = make_user('aiv_nopm', [])
        self.cv = make_client(self.v)
        self.cn = make_client(self.n)

    def test_no_permission_denied(self):
        self.assertTrue(self.cn.get('/interference/air/').json().get('error'))

    def test_view_without_add_cannot_create(self):
        resp = self.cv.post('/interference/air/',
                            data=json.dumps(air_data()),
                            content_type='application/json')
        self.assertTrue(resp.json().get('error'))
        self.assertFalse(AirInterferenceRecord.objects.exists())


class AirInterferenceCRUDTest(TestCase):
    """创建/字段校验/编辑/软删除/幂等。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('aiv_crud', ['interference.interference.view',
                                           'interference.interference.add',
                                           'interference.interference.edit',
                                           'interference.interference.del'])
        self.client = make_client(self.user)

    def _create(self, **overrides):
        return self.client.post(
            '/interference/air/',
            data=json.dumps(air_data(**overrides)),
            content_type='application/json')

    def test_create_success_and_db_state(self):
        resp = self._create()
        self.assertFalse(resp.json().get('error'), resp.json())
        record = AirInterferenceRecord.objects.get(flight_number='MU5678')
        self.assertEqual(record.handling_method, '')
        self.assertEqual(str(record.alert_altitude), '1200.00')
        self.assertEqual(record.alert_altitude_unit, 'm')
        self.assertEqual(record.duration_unit, 's')
        self.assertEqual(record.tenant_id, self.user.tenant_id)

    def test_create_required_fields(self):
        # 航班号/机型允许为空，仅日期时间与现象必填
        for field in ('datetime', 'phenomenon'):
            resp = self._create(**{field: ''})
            self.assertTrue(resp.json().get('error'), f'{field} 必填校验失败')
        # 处置方式/原因分析首次登记允许留空（不妨碍及时登记）
        resp = self._create()
        self.assertFalse(resp.json().get('error'), resp.json())

    def test_create_allows_empty_flight_number(self):
        resp = self._create(flight_number='', aircraft_type='')
        self.assertFalse(resp.json().get('error'), resp.json())
        record = AirInterferenceRecord.objects.get()
        self.assertEqual(record.flight_number, '')
        self.assertEqual(record.aircraft_type, '')

    def test_create_invalid_datetime_format(self):
        resp = self._create(datetime='2026-08-02')
        self.assertTrue(resp.json().get('error'))

    def test_create_duplicate_rejected(self):
        self._create()
        resp = self._create()
        self.assertEqual(resp.json().get('error'), '检测到重复提交，请勿重复操作')

    def test_altitude_must_be_positive(self):
        resp = self._create(alert_altitude=0)
        self.assertTrue(resp.json().get('error'))
        resp = self._create(alert_altitude=-100)
        self.assertTrue(resp.json().get('error'))
        self.assertFalse(AirInterferenceRecord.objects.exists())

    def test_altitude_unit_must_be_valid(self):
        resp = self._create(alert_altitude_unit='km')
        self.assertTrue(resp.json().get('error'))

    def test_duration_must_be_positive(self):
        resp = self._create(duration=0)
        self.assertTrue(resp.json().get('error'))
        resp = self._create(duration=-5)
        self.assertTrue(resp.json().get('error'))

    def test_duration_unit_must_be_valid(self):
        resp = self._create(duration_unit='周')
        self.assertTrue(resp.json().get('error'))

    def test_altitude_optional(self):
        resp = self._create(alert_altitude=None, duration=None)
        self.assertFalse(resp.json().get('error'), resp.json())
        record = AirInterferenceRecord.objects.get(flight_number='MU5678')
        self.assertIsNone(record.alert_altitude)
        self.assertIsNone(record.duration)

    def test_edit_updates_fields(self):
        self._create()
        record = AirInterferenceRecord.objects.get(flight_number='MU5678')
        resp = self.client.post(
            '/interference/air/',
            data=json.dumps({'id': record.id, 'handling_method': '无线电静默，雷达监控通报相关航班'},
                            ),
            content_type='application/json')
        self.assertFalse(resp.json().get('error'), resp.json())
        record.refresh_from_db()
        self.assertIn('雷达监控', record.handling_method)

    def test_edit_clearable_fields(self):
        """显式携带空串的告警高度/持续时间按清除处理；未携带的键不受影响。"""
        self._create()
        record = AirInterferenceRecord.objects.get(flight_number='MU5678')
        # 局部编辑未携带高度/持续时间：不清空
        resp = self.client.post('/interference/air/',
                                data=json.dumps({'id': record.id, 'route': 'HFE-PEK'}),
                                content_type='application/json')
        self.assertFalse(resp.json().get('error'), resp.json())
        record.refresh_from_db()
        self.assertEqual(str(record.alert_altitude), '1200.00')
        self.assertEqual(record.route, 'HFE-PEK')
        # 显式携带空串：清除
        resp = self.client.post('/interference/air/',
                                data=json.dumps({'id': record.id, 'alert_altitude': ''}),
                                content_type='application/json')
        self.assertFalse(resp.json().get('error'), resp.json())
        record.refresh_from_db()
        self.assertIsNone(record.alert_altitude)
        self.assertEqual(str(record.duration), '45.00')

    def test_soft_delete_preserves_data(self):
        self._create()
        record = AirInterferenceRecord.objects.get(flight_number='MU5678')
        resp = self.client.delete(f'/interference/air/?id={record.id}')
        self.assertFalse(resp.json().get('error'))
        raw = AirInterferenceRecord.objects.all_with_deleted().get(pk=record.id)
        self.assertTrue(raw.is_deleted)
        self.assertIsNotNone(raw.deleted_at)


class AirInterferenceTenantTest(TestCase):
    """租户隔离。"""

    def setUp(self):
        setup_test_env(self)
        self.ua = make_user('aiv_ta', ['interference.interference.view',
                                       'interference.interference.add',
                                       'interference.interference.edit',
                                       'interference.interference.del'])
        self.ua.tenant_id = 'tenant_a'
        self.ua.save()
        self.ub = make_user('aiv_tb', ['interference.interference.view',
                                       'interference.interference.add',
                                       'interference.interference.edit',
                                       'interference.interference.del'])
        self.ub.tenant_id = 'tenant_b'
        self.ub.save()
        self.ca = make_client(self.ua)
        self.cb = make_client(self.ub)
        self.ra = AirInterferenceRecord.objects.create(
            tenant_id='tenant_a',
            datetime='2026-08-02 14:30:00', flight_number='MU-A',
            phenomenon='PA', created_by=self.ua)
        self.rb = AirInterferenceRecord.objects.create(
            tenant_id='tenant_b',
            datetime='2026-08-02 14:30:00', flight_number='MU-B',
            phenomenon='PB', created_by=self.ub)

    def test_cross_tenant_list_isolated(self):
        data = self.ca.get('/interference/air/').json()['data']
        flights = [r['flight_number'] for r in data['records']]
        self.assertIn('MU-A', flights)
        self.assertNotIn('MU-B', flights)

    def test_cross_tenant_edit_blocked(self):
        resp = self.ca.post('/interference/air/',
                            data=json.dumps({'id': self.rb.id, 'phenomenon': '篡改'}),
                            content_type='application/json')
        self.assertTrue(resp.json().get('error'))
        self.rb.refresh_from_db()
        self.assertEqual(self.rb.phenomenon, 'PB')

    def test_cross_tenant_delete_blocked(self):
        resp = self.ca.delete(f'/interference/air/?id={self.rb.id}')
        self.assertTrue(resp.json().get('error'))
        self.rb.refresh_from_db()
        self.assertFalse(self.rb.is_deleted)


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class AirInterferenceAttachmentTest(TestCase):
    """附件归属隔离：空中附件不得出现在地面类型下。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('aiv_att', ['interference.interference.view',
                                          'interference.interference.add',
                                          'interference.interference.edit'])
        self.client = make_client(self.user)
        self.record = AirInterferenceRecord.objects.create(
            tenant_id=self.user.tenant_id,
            datetime='2026-08-02 14:30:00', flight_number='MU5678',
            phenomenon='P', created_by=self.user)
        self.upload_url = f'/interference/air/{self.record.id}/attachments/'
        self.file = SimpleUploadedFile('pic.png', b'\x89PNG\r\n\x1a\nbridge-air-evidence', content_type='image/png')

    def _upload(self):
        return self.client.post(self.upload_url, {'file': self.file})

    def test_upload_and_list(self):
        resp = self._upload()
        self.assertFalse(resp.json().get('error'), resp.json())
        lst = self.client.get(self.upload_url).json()['data']
        self.assertEqual(len(lst), 1)
        self.assertEqual(lst[0]['file_name'], 'pic.png')

    def test_attachment_not_visible_under_bridge_object_type(self):
        self._upload()
        # 同一记录 ID 的地面类型接口不可见该附件（跨业务类型串联被阻断）
        bridge_url = f'/interference/bridge/{self.record.id}/attachments/'
        resp = self.client.get(bridge_url).json()
        if resp.get('error'):
            # 记录不存在于地面表：返回权限错误，同样证明隔离
            pass
        else:
            self.assertEqual(resp['data'], [])
        self.assertFalse(EvidenceAttachment.objects.filter(
            module='interference', object_type='bridge_interference',
            object_id=str(self.record.id), is_deleted=False).exists())

    def test_temp_attachment_relinked_on_create(self):
        temp_id = 'temp-test-air-001'
        upload_url = f'/interference/air/{temp_id}/attachments/'
        self.client.post(upload_url, {'file': self.file})
        # 使用与 setUp 记录不同的航班号/时间，避免触发重复提交拦截
        resp = self.client.post(
            '/interference/air/',
            data=json.dumps(air_data(attachment_temp_id=temp_id,
                                     flight_number='MU_TEMP', datetime='2026-08-03 09:00:00')),
            content_type='application/json')
        self.assertFalse(resp.json().get('error'), resp.json())
        record = AirInterferenceRecord.objects.get(flight_number='MU_TEMP')
        att = EvidenceAttachment.objects.get(
            module='interference', object_type='air_interference', is_deleted=False)
        self.assertEqual(str(att.object_id), str(record.id))

    def test_cross_tenant_record_upload_blocked(self):
        other = make_user('aiv_att_b', ['interference.interference.add',
                                        'interference.interference.edit'])
        other.tenant_id = 'tenant_b'
        other.save()
        cb = make_client(other)
        resp = cb.post(self.upload_url, {'file': self.file})
        self.assertTrue(resp.json().get('error'))


class AirInterferenceExportTest(TestCase):
    """导出：列顺序符合表单字段顺序，告警高度/持续时间带单位。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('aiv_export', ['interference.interference.view',
                                             'interference.interference.add'])
        self.client = make_client(self.user)

    def test_export_columns_and_data(self):
        resp = self.client.post('/interference/air/',
                                data=json.dumps(air_data()),
                                content_type='application/json')
        self.assertFalse(resp.json().get('error'), resp.json())

        resp = self.client.get('/interference/air/export/')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])

        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.title, '空中干扰记录')
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, ['序号', '日期时间', '航班号', '机型', '航线',
                                   '被扰频率', '告警高度',
                                   '告警航段', '持续时间', '现象', '处置方式', '原因分析',
                                   '附件', '附件图片'])
        row2 = [cell.value for cell in ws[2]]
        self.assertEqual(row2[2], 'MU5678')
        self.assertEqual(row2[6], '1200米')
        self.assertEqual(row2[8], '45秒')

    def test_export_empty_data_rejected(self):
        resp = self.client.get('/interference/air/export/')
        self.assertTrue(resp.json().get('error'))


class InterferenceSummaryTest(TestCase):
    """统一汇总统计：分别统计两类并给出总量，历史记录单列。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('iv_summary', ['interference.statistics.view'])
        self.client = make_client(self.user)

    def test_summary_counts_both_types(self):
        AirInterferenceRecord.objects.create(
            tenant_id=self.user.tenant_id,
            datetime='2026-08-02 14:30:00', flight_number='MU5678',
            phenomenon='P', created_by=self.user)
        BridgeInterferenceRecord.objects.create(
            tenant_id=self.user.tenant_id,
            datetime='2026-08-01 10:00:00', flight_number='CA1234',
            phenomenon='P', created_by=self.user)
        BridgeInterferenceRecord.objects.create(
            tenant_id=self.user.tenant_id,
            datetime='2026-08-01 11:00:00', flight_number='CA1235',
            phenomenon='P', created_by=self.user)

        resp = self.client.get('/interference/summary/')
        data = resp.json()['data']
        self.assertEqual(data['bridge_count'], 2)
        self.assertEqual(data['air_count'], 1)
        self.assertEqual(data['total_count'], 3)
        # 月度趋势按记录类型分列
        by_type = {}
        for item in data['monthly_trend']:
            by_type.setdefault(item['record_type'], 0)
            by_type[item['record_type']] += item['count']
        self.assertEqual(by_type.get('bridge'), 2)
        self.assertEqual(by_type.get('air'), 1)
        # 纯记录型：不再返回状态分布
        self.assertNotIn('status_stats', data)

    def test_summary_permission_denied(self):
        user = make_user('iv_summary_nopm', [])
        client = make_client(user)
        self.assertTrue(client.get('/interference/summary/').json().get('error'))

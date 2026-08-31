# -*- coding: utf-8 -*-
"""地面无线电通信异常/干扰记录 stable_contract 测试。

覆盖：创建/字段校验/编辑/软删除/租户隔离/附件归属隔离/导出字段/权限拒绝/
状态流转。
"""
import json
import tempfile
from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings

from apps.utils.test_helpers import make_user, make_client, setup_test_env
from apps.interference.models import BridgeInterferenceRecord, AirInterferenceRecord
from apps.evidence.models import EvidenceAttachment
from apps.logs.models import AuditLog


def bridge_data(**overrides):
    data = {
        'datetime': '2026-08-01 10:00:00',
        'flight_number': 'CA1234',
        'aircraft_no': 'B-2026',
        'aircraft_type': 'A320',
        'location': 'T2航站楼3号廊桥/12号机位',
        'frequency': '118.6',
        'phenomenon': '甚高频通信出现杂音，断续无法建立联系',
        'remark': '测试备注',
    }
    data.update(overrides)
    return data


class BridgeInterferenceAuthTest(TestCase):
    """权限拒绝：未认证/无权限/缺 add 权限。"""

    def setUp(self):
        setup_test_env(self)
        self.v = make_user('biv_view', ['interference.interference.view'])
        self.adder = make_user('biv_add', ['interference.interference.view',
                                           'interference.interference.add'])
        self.n = make_user('biv_nopm', [])
        self.cv = make_client(self.v)
        self.cn = make_client(self.n)

    def test_unauthenticated_denied(self):
        from django.test import Client
        self.assertTrue(Client().get('/interference/bridge/').json().get('error'))

    def test_no_permission_denied(self):
        self.assertTrue(self.cn.get('/interference/bridge/').json().get('error'))

    def test_view_without_add_cannot_create(self):
        resp = self.cv.post('/interference/bridge/',
                            data=json.dumps(bridge_data()),
                            content_type='application/json')
        self.assertTrue(resp.json().get('error'))
        self.assertFalse(BridgeInterferenceRecord.objects.exists())

    def test_add_without_edit_cannot_edit(self):
        record = BridgeInterferenceRecord.objects.create(
            tenant_id=self.adder.tenant_id,
            datetime='2026-08-01 10:00:00', flight_number='CA0000',
            phenomenon='P', created_by=self.adder)
        resp = make_client(self.adder).post(
            '/interference/bridge/',
            data=json.dumps({'id': record.id, 'remark': '篡改'}),
            content_type='application/json')
        self.assertTrue(resp.json().get('error'))


class BridgeInterferenceCRUDTest(TestCase):
    """创建/必填校验/编辑/软删除/幂等。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('biv_crud', ['interference.interference.view',
                                           'interference.interference.add',
                                           'interference.interference.edit',
                                           'interference.interference.del'])
        self.client = make_client(self.user)

    def _create(self, **overrides):
        return self.client.post(
            '/interference/bridge/',
            data=json.dumps(bridge_data(**overrides)),
            content_type='application/json')

    def test_create_success_and_db_state(self):
        resp = self._create()
        self.assertFalse(resp.json().get('error'), resp.json())
        record = BridgeInterferenceRecord.objects.get(flight_number='CA1234')
        self.assertEqual(record.location, 'T2航站楼3号廊桥/12号机位')
        self.assertEqual(record.remark, '测试备注')
        self.assertEqual(record.tenant_id, self.user.tenant_id)

    def test_create_required_fields(self):
        # 航班号/机号/机型允许为空，仅日期时间与现象必填
        for field in ('datetime', 'phenomenon'):
            resp = self._create(**{field: ''})
            self.assertTrue(resp.json().get('error'), f'{field} 必填校验失败')
        self.assertFalse(BridgeInterferenceRecord.objects.exists())

    def test_create_allows_empty_flight_no_and_type(self):
        resp = self._create(flight_number='', aircraft_no='', aircraft_type='')
        self.assertFalse(resp.json().get('error'), resp.json())
        record = BridgeInterferenceRecord.objects.get()
        self.assertEqual(record.flight_number, '')
        self.assertEqual(record.aircraft_no, '')
        self.assertEqual(record.aircraft_type, '')

    def test_create_invalid_datetime_format(self):
        resp = self._create(datetime='2026/08/01 10:00')
        self.assertTrue(resp.json().get('error'))
        self.assertFalse(BridgeInterferenceRecord.objects.exists())

    def test_create_duplicate_rejected(self):
        self._create()
        resp = self._create()
        self.assertEqual(resp.json().get('error'), '检测到重复提交，请勿重复操作')
        self.assertEqual(BridgeInterferenceRecord.objects.count(), 1)

    def test_edit_updates_fields(self):
        self._create()
        record = BridgeInterferenceRecord.objects.get(flight_number='CA1234')
        resp = self.client.post(
            '/interference/bridge/',
            data=json.dumps({'id': record.id, 'location': '15号机位', 'remark': ''}),
            content_type='application/json')
        self.assertFalse(resp.json().get('error'), resp.json())
        record.refresh_from_db()
        self.assertEqual(record.location, '15号机位')
        self.assertIsNotNone(record.updated_at)

    def test_edit_nonexistent_blocked(self):
        resp = self.client.post(
            '/interference/bridge/',
            data=json.dumps({'id': 99999, 'remark': 'x'}),
            content_type='application/json')
        self.assertTrue(resp.json().get('error'))

    def test_soft_delete_preserves_data(self):
        self._create()
        record = BridgeInterferenceRecord.objects.get(flight_number='CA1234')
        resp = self.client.delete(f'/interference/bridge/?id={record.id}')
        self.assertFalse(resp.json().get('error'))
        raw = BridgeInterferenceRecord.objects.all_with_deleted().get(pk=record.id)
        self.assertTrue(raw.is_deleted)
        self.assertIsNotNone(raw.deleted_at)
        # 列表不再可见
        resp = self.client.get('/interference/bridge/')
        self.assertEqual(resp.json()['data']['total'], 0)

    def test_audit_log_records_create_and_delete(self):
        self._create()
        self.assertTrue(AuditLog.objects.filter(action='create').exists())
        record = BridgeInterferenceRecord.objects.get(flight_number='CA1234')
        self.client.delete(f'/interference/bridge/?id={record.id}')
        self.assertTrue(AuditLog.objects.filter(action='delete').exists())


class BridgeInterferenceTenantTest(TestCase):
    """租户隔离：列表/编辑/删除/附件。"""

    def setUp(self):
        setup_test_env(self)
        self.ua = make_user('biv_ta', ['interference.interference.view',
                                       'interference.interference.add',
                                       'interference.interference.edit',
                                       'interference.interference.del'])
        self.ua.tenant_id = 'tenant_a'
        self.ua.save()
        self.ub = make_user('biv_tb', ['interference.interference.view',
                                       'interference.interference.add',
                                       'interference.interference.edit',
                                       'interference.interference.del'])
        self.ub.tenant_id = 'tenant_b'
        self.ub.save()
        self.ca = make_client(self.ua)
        self.cb = make_client(self.ub)
        self.ra = BridgeInterferenceRecord.objects.create(
            tenant_id='tenant_a',
            datetime='2026-08-01 10:00:00', flight_number='CA-A',
            phenomenon='PA', created_by=self.ua)
        self.rb = BridgeInterferenceRecord.objects.create(
            tenant_id='tenant_b',
            datetime='2026-08-01 10:00:00', flight_number='CA-B',
            phenomenon='PB', created_by=self.ub)

    def test_cross_tenant_list_isolated(self):
        data = self.ca.get('/interference/bridge/').json()['data']
        flights = [r['flight_number'] for r in data['records']]
        self.assertIn('CA-A', flights)
        self.assertNotIn('CA-B', flights)

    def test_cross_tenant_edit_blocked(self):
        resp = self.ca.post('/interference/bridge/',
                            data=json.dumps({'id': self.rb.id, 'phenomenon': '篡改'}),
                            content_type='application/json')
        self.assertTrue(resp.json().get('error'))
        self.rb.refresh_from_db()
        self.assertEqual(self.rb.phenomenon, 'PB')

    def test_cross_tenant_delete_blocked(self):
        resp = self.ca.delete(f'/interference/bridge/?id={self.rb.id}')
        self.assertTrue(resp.json().get('error'))
        self.rb.refresh_from_db()
        self.assertFalse(self.rb.is_deleted)


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class BridgeInterferenceAttachmentTest(TestCase):
    """附件归属隔离：地面附件不得出现在空中类型下；跨租户不可见。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('biv_att', ['interference.interference.view',
                                          'interference.interference.add',
                                          'interference.interference.edit'])
        self.other = make_user('biv_att_b', ['interference.interference.view'])
        self.other.tenant_id = 'tenant_b'
        self.other.save()
        self.client = make_client(self.user)
        self.cb = make_client(self.other)
        self.record = BridgeInterferenceRecord.objects.create(
            tenant_id=self.user.tenant_id,
            datetime='2026-08-01 10:00:00', flight_number='CA1234',
            phenomenon='P', created_by=self.user)
        self.upload_url = f'/interference/bridge/{self.record.id}/attachments/'
        self.file = SimpleUploadedFile('pic.png', b'\x89PNG\r\n\x1a\nbridge-air-evidence', content_type='image/png')

    def test_upload_and_list(self):
        resp = self.client.post(self.upload_url, {'file': self.file})
        self.assertFalse(resp.json().get('error'), resp.json())
        lst = self.client.get(self.upload_url).json()['data']
        self.assertEqual(len(lst), 1)
        self.assertEqual(lst[0]['file_name'], 'pic.png')

    def test_upload_requires_file(self):
        resp = self.client.post(self.upload_url, {})
        self.assertTrue(resp.json().get('error'))

    def test_attachment_not_visible_under_air_object_type(self):
        """object_type 隔离：地面附件不得通过空中业务类型查到。"""
        self.client.post(self.upload_url, {'file': self.file})
        air_url = f'/interference/air/{self.record.id}/attachments/'
        resp = self.client.get(air_url).json()
        if resp.get('error'):
            # 记录不存在于空中表：返回权限错误，同样证明隔离
            pass
        else:
            self.assertEqual(resp['data'], [])
        # 且空中类型下确实没有该 object_id 的附件
        self.assertFalse(EvidenceAttachment.objects.filter(
            module='interference', object_type='air_interference',
            object_id=str(self.record.id), is_deleted=False).exists())

    def test_cross_tenant_attachment_list_blocked(self):
        self.client.post(self.upload_url, {'file': self.file})
        resp = self.cb.get(self.upload_url)
        self.assertTrue(resp.json().get('error'))

    def test_temp_attachment_relinked_on_create(self):
        """新建未保存记录的临时附件在保存后关联到新记录。"""
        temp_id = 'temp-test-bridge-001'
        upload_url = f'/interference/bridge/{temp_id}/attachments/'
        self.client.post(upload_url, {'file': self.file})
        self.assertTrue(EvidenceAttachment.objects.filter(
            module='interference', object_type='bridge_interference',
            object_id=temp_id, is_deleted=False).exists())
        # 使用与 setUp 记录不同的航班号/时间，避免触发重复提交拦截
        resp = self.client.post(
            '/interference/bridge/',
            data=json.dumps(bridge_data(attachment_temp_id=temp_id,
                                        flight_number='CA_TEMP', datetime='2026-08-03 09:00:00')),
            content_type='application/json')
        self.assertFalse(resp.json().get('error'), resp.json())
        record = BridgeInterferenceRecord.objects.get(flight_number='CA_TEMP')
        att = EvidenceAttachment.objects.get(
            module='interference', object_type='bridge_interference', is_deleted=False)
        self.assertEqual(str(att.object_id), str(record.id))

    def test_download(self):
        self.client.post(self.upload_url, {'file': self.file})
        att = EvidenceAttachment.objects.get(
            module='interference', object_type='bridge_interference', is_deleted=False)
        resp = self.client.get(f'/interference/attachments/{att.id}/download/')
        self.assertEqual(resp.status_code, 200)
        # FileResponse 为流式响应，聚合流内容校验文件字节
        content = b''.join(resp.streaming_content)
        self.assertTrue(content.startswith(b'\x89PNG'))


class BridgeInterferenceExportTest(TestCase):
    """导出：列顺序符合表单字段顺序，数据落列正确。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('biv_export', ['interference.interference.view',
                                             'interference.interference.add'])
        self.client = make_client(self.user)

    def test_export_columns_and_data(self):
        resp = self.client.post(
            '/interference/bridge/',
            data=json.dumps(bridge_data()),
            content_type='application/json')
        self.assertFalse(resp.json().get('error'), resp.json())

        resp = self.client.get('/interference/bridge/export/')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])

        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, ['序号', '日期时间', '航班号', '机号', '机型',
                                   '位置/机位', '频率', '现象', '备注', '附件', '附件图片'])
        row2 = [cell.value for cell in ws[2]]
        self.assertEqual(row2[2], 'CA1234')
        self.assertEqual(row2[5], 'T2航站楼3号廊桥/12号机位')

    def test_export_empty_data_rejected(self):
        resp = self.client.get('/interference/bridge/export/')
        self.assertTrue(resp.json().get('error'))


class BridgeAirFieldBoundaryTest(TestCase):
    """业务边界：地面记录不承载空中字段，空中记录不承载地面字段。"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('biv_boundary', ['interference.interference.view'])
        self.client = make_client(self.user)

    def test_bridge_model_has_no_air_fields(self):
        names = {f.name for f in BridgeInterferenceRecord._meta.get_fields()}
        for air_field in ('route', 'runway', 'approach_procedure', 'alert_form',
                          'alert_altitude', 'duration', 'handling_method', 'cause_analysis'):
            self.assertNotIn(air_field, names)

    def test_air_model_has_no_bridge_fields(self):
        names = {f.name for f in AirInterferenceRecord._meta.get_fields()}
        for bridge_field in ('location', 'frequency', 'aircraft_no', 'remark'):
            self.assertNotIn(bridge_field, names)

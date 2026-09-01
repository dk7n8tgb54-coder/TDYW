# Copyright: (c) OpenSpug Organization. https://github.com/openspug
# Copyright (c) <spug.dev@gmail.com>
# Released under the AGPL-3.0 License.
"""干扰管理双业务类型 Excel 导入服务（地面/空中）。

职责边界：
- 生成两类业务各自的 .xlsx 导入模板（标题 + 填写说明 + 文本格式表头）；
- 解析上传的 Excel：在前 20 行内自动定位表头行，按表头名称映射字段
  （仅支持有限、明确的别名，不猜测模糊名称），从表头下一行读取连续数据，
  遇到第一行完全空白行即停止，避免把页脚说明误导入为数据；
- 行级校验复用 business_views 的手工表单校验逻辑（必填/日期格式/正数/
  单位合法性），并按模型字段 max_length 兜底，Excel 导入不绕过任何业务规则；
- 生成错误报告（原文件列 + Excel行号 + 错误原因）。

安全约束：
- 仅支持 .xlsx（.xls 在视图层明确提示另存为 .xlsx）；
- 拒绝数据单元格中的公式；写入模板与错误报告时对 =, +, -, @ 开头的文本
  强制按字符串存储（data_type='s'），防止公式注入；
- 数据表头和数据行不允许合并单元格（标题与说明区允许）；一个表头单元格
  用斜线表达两个字段（如「航班号/机号」）时拒绝导入并提示拆分为两列；
- 文件大小与数据行数上限（settings 可覆盖），超限 fail-closed；
- 本服务不写数据库，业务写入由 import_views 在事务内完成。

业务边界：
- 业务类型由上传入口（地面/空中页面各自的接口）决定，不从文件内容猜测；
- 上传文件中出现另一业务类型的专用表头时直接拒绝，防止数据被静默丢弃。
"""
import io
import logging
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.db.models import Q
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from libs.tenant_utils import apply_tenant_filter

from apps.interference.business_views import (
    AIR_OBJECT_TYPE,
    BRIDGE_OBJECT_TYPE,
    AirInterferenceView,
    BridgeInterferenceView,
    _validate_positive_number,
)
from apps.interference.models import (
    AirInterferenceRecord,
    BridgeInterferenceRecord,
    format_decimal,
)

logger = logging.getLogger(__name__)

# ==== 导入限制（fail-closed，可通过 Django settings 覆盖） ====
IMPORT_MAX_HEADER_SCAN_ROWS = 20
IMPORT_TEXT_CELL_ROWS = 200  # 模板中预置文本格式的数据行数


def import_max_file_mb():
    """导入文件大小上限（MB）。"""
    return int(getattr(settings, 'INTERFERENCE_IMPORT_MAX_FILE_MB', 10))


def import_max_rows():
    """单次导入的数据行数上限。"""
    return int(getattr(settings, 'INTERFERENCE_IMPORT_MAX_ROWS', 2000))


class ImportParseError(Exception):
    """文件级结构性错误（表头缺失/重复/合并单元格/跨业务文件等），终止导入。"""


# ==== 表头字段定义 ====
# kind: None=文本字段；datetime=日期时间；altitude/altitude_unit/duration=空中专用
BRIDGE_IMPORT_FIELDS = [
    {'key': 'datetime', 'label': '日期时间', 'required': True, 'kind': 'datetime',
     'aliases': ('发生时间',)},
    # 航班号/机号/机型均允许为空
    {'key': 'flight_number', 'label': '航班号', 'aliases': ('航班',)},
    {'key': 'aircraft_no', 'label': '机号', 'aliases': ('飞机号', '航空器注册号')},
    {'key': 'aircraft_type', 'label': '机型'},
    # 位置与机位为同一字段（位置/机位）；「位置」「机位」作为兼容别名，
    # 旧格式文件仅含其中一列时仍可映射导入，同时出现两列将按重复表头拒绝
    {'key': 'location', 'label': '位置/机位',
     'aliases': ('位置', '地点', '机位', '机位号')},
    {'key': 'frequency', 'label': '频率', 'aliases': ('工作频率',)},
    {'key': 'phenomenon', 'label': '现象', 'required': True, 'aliases': ('干扰现象',)},
    {'key': 'handling_method', 'label': '处置方式', 'aliases': ('处置措施',)},
    {'key': 'cause_analysis', 'label': '原因分析', 'aliases': ('原因',)},
    {'key': 'remark', 'label': '备注'},
]

AIR_IMPORT_FIELDS = [
    {'key': 'datetime', 'label': '日期时间', 'required': True, 'kind': 'datetime',
     'aliases': ('发生时间',)},
    {'key': 'flight_number', 'label': '航班号', 'aliases': ('航班',)},
    {'key': 'aircraft_type', 'label': '机型'},
    {'key': 'route', 'label': '航线', 'aliases': ('航路',)},
    {'key': 'alert_form', 'label': '被扰频率', 'aliases': ('告警形式', '告警类型')},
    {'key': 'alert_altitude', 'label': '告警高度', 'kind': 'altitude'},
    {'key': 'alert_altitude_unit', 'label': '高度单位', 'kind': 'altitude_unit',
     'aliases': ('告警高度单位',)},
    {'key': 'alert_segment', 'label': '告警航段', 'aliases': ('航段',)},
    {'key': 'duration', 'label': '持续时间', 'kind': 'duration', 'aliases': ('时长',)},
    {'key': 'phenomenon', 'label': '现象', 'required': True, 'aliases': ('干扰现象',)},
    {'key': 'handling_method', 'label': '处置方式', 'aliases': ('处置措施',)},
    {'key': 'cause_analysis', 'label': '原因分析', 'aliases': ('原因',)},
]

# 高度单位的明确别名映射（不换算数值，仅归一化单位文本）
ALTITUDE_UNIT_ALIASES = {'米': 'm', 'm': 'm', '英尺': 'ft', 'ft': 'ft'}


def _norm_text(value):
    """表头归一化：转文本、去首尾空白并压缩内部连续空白。"""
    if value is None:
        return ''
    return ' '.join(str(value).split())


COMMON_TEMPLATE_NOTES = [
    '填写说明：',
    '1. 表头行（下方首个字段行）名称请勿修改；「日期时间、航班号、现象」为必填列。',
    '2. 日期时间：YYYY-MM-DD HH:MM（到分钟即可，如 2026-08-01 10:30，2026.7.8 17:15 这类写法也可）；也可仅填日期（按 00:00 处理）或使用 Excel 日期单元格；分隔符支持 - / . 与中文；航班号、机号、机型允许留空。',
    '3. 航班号、机号、机位等列已设置为文本格式，可直接输入 001 等带前导零的内容，不会丢失前导零。',
    '4. 数据区请勿使用合并单元格和公式；每列只表达一个字段，例如「航班号」「机号」请分成两列，不要写成「航班号/机号」。',
    '5. 未识别的列不会导入并在预校验时提示；数据区出现第一个空行后，其后内容不会被导入。',
]

BUSINESS_CONFIG = {
    'bridge': {
        'key': 'bridge',
        'label': '地面无线电通信异常/干扰',
        'model': BridgeInterferenceRecord,
        'view': BridgeInterferenceView,
        'object_type': BRIDGE_OBJECT_TYPE,
        'fields': BRIDGE_IMPORT_FIELDS,
        'template_title': '地面无线电通信异常/干扰记录导入模板',
        'template_filename': '地面干扰导入模板',
        # 文件内/数据库重复判断键（仅用于预警与拒绝，不建立数据库唯一约束）
        'duplicate_fields': ('datetime', 'flight_number', 'location', 'frequency'),
        'duplicate_label': '日期时间 + 航班号 + 位置/机位 + 频率',
        'notes': COMMON_TEMPLATE_NOTES + [
            '6. 处置方式、原因分析在首次登记时允许留空。',
        ],
    },
    'air': {
        'key': 'air',
        'label': '空中干扰',
        'model': AirInterferenceRecord,
        'view': AirInterferenceView,
        'object_type': AIR_OBJECT_TYPE,
        'fields': AIR_IMPORT_FIELDS,
        'template_title': '空中干扰记录导入模板',
        'template_filename': '空中干扰导入模板',
        'duplicate_fields': ('datetime', 'flight_number', 'alert_form', 'alert_segment'),
        'duplicate_label': '日期时间 + 航班号 + 被扰频率 + 告警航段',
        'notes': COMMON_TEMPLATE_NOTES + [
            '6. 高度单位：只能填写「米」或「英尺」，与告警高度分别填写，系统不做单位换算。',
            '7. 持续时间：HH:MM:SS 格式（如 01:30:00 表示 1 小时 30 分钟）。',
            '8. 处置方式、原因分析在首次登记时允许留空。',
        ],
    },
}


def _build_alias_map(field_specs):
    """构建 归一化表头文本 -> 字段key 的精确匹配映射（含有限别名）。"""
    mapping = {}
    for field in field_specs:
        names = {field['label'], *field.get('aliases', ())}
        for name in names:
            mapping.setdefault(_norm_text(name), field['key'])
    return mapping


for _config in BUSINESS_CONFIG.values():
    _config['alias_map'] = _build_alias_map(_config['fields'])


def _cell_to_text(value):
    """单元格值统一转文本：日期归一化到分钟，数字去掉多余小数尾零。"""
    if value is None:
        return ''
    if isinstance(value, datetime):
        if value.microsecond:
            # Excel 日期串行值的浮点噪声（如 09:59:59.999999）：先按最近整秒归位
            if value.microsecond >= 500000:
                value = value + timedelta(seconds=1)
            value = value.replace(microsecond=0)
        # 业务精度到分钟，秒截断为 00
        return value.strftime('%Y-%m-%d %H:%M')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, bool):
        return '是' if value else '否'
    if isinstance(value, (int, float)):
        return format_decimal(value)
    return str(value).strip()


def _safe_cell(ws, row, col, value):
    """写入单元格；对以 =, +, -, @ 开头的文本强制按字符串存储，防止公式注入。"""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if isinstance(value, str) and value[:1] in ('=', '+', '-', '@'):
        cell.data_type = 's'
    return cell


# ====================================================================
# 模板生成
# ====================================================================

def _write_template_header_cell(ws, row, col, field):
    cell = ws.cell(row=row, column=col, value=field['label'])
    cell.font = Font(name='宋体', bold=True, color='FFFFFF', size=11)
    if field.get('required'):
        cell.font = Font(name='宋体', bold=True, color='FF0000', size=11)
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='BFBFBF')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell


def build_template_workbook(business):
    """构建业务导入模板（.xlsx），返回 BytesIO。"""
    config = BUSINESS_CONFIG[business]
    fields = config['fields']
    ncols = len(fields)

    wb = Workbook()
    ws = wb.active
    ws.title = '导入模板'

    # 标题与说明区：允许合并单元格；说明行整行合并便于阅读
    title_cell = ws.cell(row=1, column=1, value=config['template_title'])
    title_cell.font = Font(name='宋体', bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

    note_start_row = 3
    for idx, note in enumerate(config['notes']):
        row = note_start_row + idx
        note_cell = ws.cell(row=row, column=1, value=note)
        note_cell.font = Font(name='宋体', size=10, color='595959')
        note_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

    header_row = note_start_row + len(config['notes']) + 1
    for col_idx, field in enumerate(fields, start=1):
        _write_template_header_cell(ws, header_row, col_idx, field)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(field['label']) * 2 + 4, 12)

    # 数据区前 IMPORT_TEXT_CELL_ROWS 行全部列设为文本格式，避免前导零丢失
    for row in range(header_row + 1, header_row + 1 + IMPORT_TEXT_CELL_ROWS):
        for col_idx in range(1, ncols + 1):
            ws.cell(row=row, column=col_idx).number_format = '@'

    ws.freeze_panes = f'A{header_row + 1}'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ====================================================================
# Excel 解析与校验
# ====================================================================

def _detect_header_row(ws):
    """在前 IMPORT_MAX_HEADER_SCAN_ROWS 行内定位表头行：首个命中>=2个已知表头名的行。"""
    for row_idx in range(1, min(ws.max_row, IMPORT_MAX_HEADER_SCAN_ROWS) + 1):
        matched = sum(
            1 for cell in ws[row_idx] if _norm_text(cell.value) and _norm_text(cell.value) in _all_known_names())
        if matched >= 2:
            return row_idx
    raise ImportParseError(
        f'前 {IMPORT_MAX_HEADER_SCAN_ROWS} 行内未找到有效的表头行，请下载导入模板并按模板格式填写')


def _all_known_names():
    """两类业务全部已知表头名（用于表头行定位）。"""
    names = set()
    for config in BUSINESS_CONFIG.values():
        names.update(config['alias_map'].keys())
    return names


def _map_header_columns(ws, header_row, config, other_config):
    """按表头名称映射列；返回 (col_map, unknown_cols, header_labels)。

    结构性错误直接抛出 ImportParseError：
    - 重复表头（同一字段对应多列）；
    - 斜线组合表头（一个单元格表达两个字段，如「航班号/机号」）；
    - 另一业务类型的专用表头（防止跨业务上传静默丢列）；
    - 必需表头缺失；
    - 表头行存在合并单元格。
    """
    name_to_key = config['alias_map']
    other_names = other_config['alias_map']
    required_labels = {f['label'] for f in config['fields'] if f.get('required')}

    col_map = {}
    unknown_cols = []
    header_labels = {}
    duplicate_hits = []
    slash_hits = []
    cross_hits = []

    for cell in ws[header_row]:
        col_idx = cell.column
        label = _norm_text(cell.value)
        if not label:
            continue
        header_labels[col_idx] = str(cell.value).strip()
        key = name_to_key.get(label)
        if key:
            if key in col_map:
                duplicate_hits.append((key, col_idx, col_map[key]))
            else:
                col_map[key] = col_idx
            continue
        # 斜线组合表头：按 / 或 ／ 拆分后各部分均为已知字段名
        parts = [part for part in re.split(r'[/／]', label) if part]
        if len(parts) >= 2 and all(part in name_to_key for part in parts):
            slash_hits.append((label, [name_to_key[part] for part in parts]))
        elif len(parts) >= 2 and all(part in other_names for part in parts):
            cross_hits.append(label)
        elif label in other_names:
            cross_hits.append(label)
        else:
            unknown_cols.append((col_idx, str(cell.value).strip()))

    # 表头行合并单元格：标题/说明区允许合并，表头行不允许
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= header_row <= merged_range.max_row:
            raise ImportParseError(
                f'表头行（第 {header_row} 行）存在合并单元格（{merged_range.coord}），'
                '请取消合并后重新导入')

    if duplicate_hits:
        key, col_b, col_a = duplicate_hits[0]
        label = _key_label(config, key)
        raise ImportParseError(
            f'存在重复表头：「{label}」（第 {get_column_letter(col_a)} 列与'
            f'第 {get_column_letter(col_b)} 列），请删除多余列后重新导入')

    if slash_hits:
        label, keys = slash_hits[0]
        labels = '、'.join(_key_label(config, k) for k in keys)
        raise ImportParseError(
            f'表头「{label}」将多个字段合并在一个单元格中（{labels}），'
            '请拆分为独立的两列后重新导入')

    if cross_hits:
        raise ImportParseError(
            f'检测到「{other_config["label"]}」业务专用的表头（{cross_hits[0]}）：'
            f'本入口仅导入「{config["label"]}」记录，请在对应业务页面导入该文件')

    missing = [label for label in required_labels
               if _key_by_label(config, label) not in col_map]
    if missing:
        raise ImportParseError(f'缺少必需表头：{"、".join(missing)}，请按导入模板补齐后重新导入')

    return col_map, unknown_cols, header_labels


def _key_label(config, key):
    for field in config['fields']:
        if field['key'] == key:
            return field['label']
    return key


def _key_by_label(config, label):
    for field in config['fields']:
        if field['label'] == label:
            return field['key']
    return label


def _check_data_merged_cells(ws, header_row, data_rows):
    """数据行合并单元格检查：返回 row -> 合并区域描述列表。"""
    data_row_set = {row['excel_row'] for row in data_rows}
    merged_by_row = {}
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row > header_row and merged_range.min_row in data_row_set:
            merged_by_row.setdefault(merged_range.min_row, []).append(merged_range.coord)
    return merged_by_row


# 日期时间通用写法：YYYY-MM-DD / YYYY/M/D / YYYY.M.D / YYYY年M月D日，
# 时分可选（HH:MM 或 HH:MM:SS，秒忽略），日期与时间之间允许任意空白
DATETIME_RE = re.compile(
    r'^(\d{4})[-/年.](\d{1,2})[-/月.](\d{1,2})日?'
    r'(?:\s+(\d{1,2}):(\d{2})(?::(\d{2}))?)?$')

DATETIME_IMPORT_ERROR = ('日期时间无法识别，请填写如 2026-08-01 10:30（到分钟即可）；'
                         '也可仅填日期（2026-08-01、2026/8/1、2026.8.1、2026年8月1日 均可）'
                         '或使用 Excel 日期单元格')


def _clean_datetime(raw_text):
    """清洗日期时间：到分钟即可（秒忽略）。

    支持（分隔符 - / . 中文 均可）：
    - 仅日期：2026-08-01、2026/8/1、2026.8.1、2026年8月1日（按 00:00 处理）；
    - 日期 + 时分/时分秒：2024.7.8 17:15、2026-08-01 10:30:45（秒忽略）；
    - Excel 日期单元格（仅日期按 00:00 处理）。
    统一归一化为分钟精度 YYYY-MM-DD HH:MM。
    """
    # 压缩任意空白（含 Excel 单元格内换行）为单个空格
    text = ' '.join(str(raw_text or '').split())
    if not text:
        return '', None
    match = DATETIME_RE.match(text)
    if match:
        year, month, day = (int(match.group(i)) for i in (1, 2, 3))
        hour = int(match.group(4)) if match.group(4) else 0
        minute = int(match.group(5)) if match.group(5) else 0
        # 秒（group 6）业务精度到分钟，忽略
        try:
            parsed = datetime(year, month, day, hour, minute)
        except ValueError:
            return raw_text, DATETIME_IMPORT_ERROR
        return parsed.strftime('%Y-%m-%d %H:%M'), None
    return raw_text, DATETIME_IMPORT_ERROR


def _clean_row(config, raw, excel_row, errors, warnings):
    """行级清洗与校验：完全复用手工表单的后端校验逻辑，返回 cleaned dict。"""
    cleaned = {}
    model = config['model']
    view = config['view']()

    # 必填字段（与手工表单 required_fields 一致，错误文案一致）
    for field_key, label in view.required_fields.items():
        if not (raw.get(field_key) or '').strip():
            errors.append(_row_error(excel_row, _key_label(config, field_key),
                                     f'请输入{label}', raw.get(field_key, '')))

    # 日期时间
    datetime_text, datetime_error = _clean_datetime(raw.get('datetime'))
    if datetime_error:
        errors.append(_row_error(excel_row, '日期时间', datetime_error, raw.get('datetime', '')))
    elif datetime_text:
        cleaned['datetime'] = datetime_text

    # 文本字段：长度按模型 max_length 兜底（手工表单由数据库约束兜底）
    for field in config['fields']:
        key = field['key']
        if field.get('kind'):
            continue
        value = (raw.get(key) or '').strip()
        if not value:
            continue
        max_length = _field_max_length(model, key)
        if max_length and len(value) > max_length:
            errors.append(_row_error(excel_row, field['label'],
                                     f'「{field["label"]}」长度不能超过{max_length}个字符',
                                     _summary(value)))
            continue
        cleaned[key] = value

    view_values = {}
    if config['key'] == 'air':
        _clean_air_numeric(config, raw, excel_row, cleaned, view_values, errors, warnings)

    # 数值/单位合法性兜底校验（与手工表单 _validate_business 同一实现）
    if view_values:
        error = view._validate_business(view_values)
        if error:
            errors.append(_row_error(excel_row, '', error, ''))

    return cleaned


def _field_max_length(model, key):
    try:
        return model._meta.get_field(key).max_length
    except Exception:
        return None


def _clean_air_numeric(config, raw, excel_row, cleaned, view_values, errors, warnings):
    """空中专用数值字段：告警高度+高度单位（不换算）、持续时间（HH:MM:SS）。"""
    altitude_raw = (raw.get('alert_altitude') or '').strip()
    unit_raw = (raw.get('alert_altitude_unit') or '').strip()

    unit = None
    if unit_raw:
        unit = ALTITUDE_UNIT_ALIASES.get(unit_raw.lower())
        if unit is None:
            errors.append(_row_error(excel_row, '高度单位', '「高度单位」只能填写「米」或「英尺」',
                                     _summary(unit_raw)))
    if altitude_raw:
        if unit is None:
            errors.append(_row_error(
                excel_row, '高度单位',
                '填写了「告警高度」时必须填写「高度单位」（米 或 英尺）', ''))
        else:
            error = _validate_positive_number({'alert_altitude': altitude_raw},
                                              'alert_altitude', '告警高度')
            if error:
                errors.append(_row_error(excel_row, '告警高度', error, _summary(altitude_raw)))
            else:
                try:
                    altitude = Decimal(altitude_raw)
                except InvalidOperation:
                    errors.append(_row_error(excel_row, '告警高度', '告警高度必须为数字',
                                             _summary(altitude_raw)))
                else:
                    # 高度与单位分别保存原始值，不做换算
                    cleaned['alert_altitude'] = altitude
                    cleaned['alert_altitude_unit'] = unit
                    view_values['alert_altitude'] = altitude
                    view_values['alert_altitude_unit'] = unit
    elif unit_raw:
        warnings.append(_row_warning(
            excel_row, '高度单位', '已填写「高度单位」但未填写「告警高度」，该单位不会保存',
            _summary(unit_raw)))

    duration_raw = (raw.get('duration') or '').strip()
    if duration_raw:
        match = re.fullmatch(r'(\d+):([0-5]?\d):([0-5]?\d)', duration_raw)
        if not match:
            errors.append(_row_error(
                excel_row, '持续时间',
                '「持续时间」必须为 HH:MM:SS 格式（如 01:30:00 表示 1 小时 30 分钟）',
                _summary(duration_raw)))
        else:
            hours, minutes, seconds = (int(part) for part in match.groups())
            total_seconds = hours * 3600 + minutes * 60 + seconds
            if total_seconds <= 0:
                errors.append(_row_error(excel_row, '持续时间', '「持续时间」必须大于0',
                                         _summary(duration_raw)))
            else:
                # 统一换算为秒保存（HH:MM:SS -> 秒，无损），单位固定为 s
                cleaned['duration'] = Decimal(total_seconds)
                cleaned['duration_unit'] = 's'
                view_values['duration'] = Decimal(total_seconds)
                view_values['duration_unit'] = 's'


def _row_error(excel_row, field, message, value):
    return {'row': excel_row, 'field': field, 'message': message, 'value': str(value or '')[:100]}


def _row_warning(excel_row, field, message, value):
    return {'row': excel_row, 'field': field, 'message': message, 'value': str(value or '')[:100]}


def _summary(value, limit=50):
    text = str(value or '')
    return text if len(text) <= limit else text[:limit] + '…'


def _duplicate_key(config, cleaned):
    return tuple((cleaned.get(key) or '') for key in config['duplicate_fields'])


def _minute_window(text):
    """分钟精度时间窗 [text, text+1min)：用于同分钟记录的数据库重复匹配
    （遵守 DateTimeField 禁用 __date/__year 等查找的规则，改用 gte/lt）。"""
    start = datetime.strptime(text, '%Y-%m-%d %H:%M')
    end = start + timedelta(minutes=1)
    return start.strftime('%Y-%m-%d %H:%M:%S'), end.strftime('%Y-%m-%d %H:%M:%S')


def find_db_duplicate_rows(business, user, rows):
    """数据库重复预警：按业务重复键（租户隔离）查询既有记录。

    仅用于拒绝重复导入，不建立数据库唯一约束。
    日期时间按分钟窗口匹配（业务精度到分钟，历史秒级记录同分钟亦视为重复）。
    返回 {excel_row: 重复描述}。
    """
    config = BUSINESS_CONFIG[business]
    model = config['model']
    key_fields = config['duplicate_fields']

    conditions = Q()
    has_condition = False
    row_keys = {}
    for row in rows:
        if row['errors']:
            continue
        cleaned = row['cleaned']
        if not cleaned.get('datetime'):
            continue
        key = _duplicate_key(config, cleaned)
        row_keys[row['excel_row']] = key
        query_kwargs = {}
        for idx, field in enumerate(key_fields):
            if field == 'datetime':
                start, end = _minute_window(key[idx])
                query_kwargs['datetime__gte'] = start
                query_kwargs['datetime__lt'] = end
            else:
                query_kwargs[field] = key[idx]
        conditions |= Q(**query_kwargs)
        has_condition = True
    if not has_condition:
        return {}

    existing_keys = set()
    qs = apply_tenant_filter(model.objects.filter(is_deleted=False), user)
    for values in qs.filter(conditions).values_list(*key_fields):
        normalized = []
        for value in values:
            if isinstance(value, datetime):
                # 与行键同精度：分钟
                normalized.append(value.strftime('%Y-%m-%d %H:%M'))
            else:
                normalized.append(value or '')
        existing_keys.add(tuple(normalized))

    duplicates = {}
    for excel_row, key in row_keys.items():
        if key in existing_keys:
            duplicates[excel_row] = (
                f'数据库中已存在相同记录（判断依据：{config["duplicate_label"]}），'
                '默认拒绝重复导入，请勿重复登记')
    return duplicates


def parse_workbook(business, data):
    """解析并校验上传的 Excel（不写数据库）。

    返回 dict：
    - header_row/header_labels/total_rows/stats/errors/warnings/rows
    - rows[i] = {excel_row, cleaned, errors, warnings, raw_by_col}
    文件级结构性错误抛出 ImportParseError。
    """
    config = BUSINESS_CONFIG[business]
    other_config = BUSINESS_CONFIG['air' if business == 'bridge' else 'bridge']

    try:
        workbook = load_workbook(io.BytesIO(data))
    except Exception:
        logger.warning('[InterferenceImport] Excel 解析失败 business=%s', business, exc_info=True)
        raise ImportParseError('无法读取该 Excel 文件，请确认文件未损坏且为 .xlsx 格式')
    worksheet = workbook.active
    if worksheet is None:
        raise ImportParseError('Excel 文件中没有工作表')

    header_row = _detect_header_row(worksheet)
    col_map, unknown_cols, header_labels = _map_header_columns(
        worksheet, header_row, config, other_config)

    warnings = []
    for col_idx, label in unknown_cols:
        warnings.append({
            'row': None, 'field': '',
            'message': f'未识别的列「{label}」（第 {get_column_letter(col_idx)} 列），该列内容不会导入',
            'value': '',
        })

    # 读取连续数据行：第一行完全空白行即停止
    rows = []
    max_rows = import_max_rows()
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        if all(cell.value in (None, '') for cell in worksheet[row_idx]):
            break
        rows.append(row_idx)
        if len(rows) > max_rows:
            raise ImportParseError(
                f'数据行数超过上限（最多 {max_rows} 行），请拆分文件后分批导入')

    merged_by_row = _check_data_merged_cells(worksheet, header_row, [
        {'excel_row': row_idx} for row_idx in rows])

    parsed_rows = []
    for row_idx in rows:
        row_errors = []
        row_warnings = []
        if row_idx in merged_by_row:
            coords = '、'.join(merged_by_row[row_idx])
            row_errors.append(_row_error(
                row_idx, '', f'第 {row_idx} 行存在合并单元格（{coords}），请取消合并后重新导入', ''))
            parsed_rows.append({
                'excel_row': row_idx, 'cleaned': {}, 'errors': row_errors,
                'warnings': row_warnings, 'raw_by_col': {},
            })
            continue

        raw = {}
        raw_by_col = {}
        for key, col_idx in col_map.items():
            cell = worksheet.cell(row=row_idx, column=col_idx)
            label = _key_label(config, key)
            if cell.data_type == 'f':
                row_errors.append(_row_error(
                    row_idx, label, f'「{label}」列不能使用公式，请以纯文本数值填写',
                    _summary(cell.value)))
                raw[key] = ''
                continue
            raw[key] = _cell_to_text(cell.value)
        # 未识别列仅用于错误报告回显，不参与导入
        for col_idx, _label in unknown_cols:
            raw_by_col[col_idx] = _cell_to_text(worksheet.cell(row=row_idx, column=col_idx).value)
        raw_by_col.update({col_map[key]: raw.get(key, '') for key in col_map})

        cleaned = _clean_row(config, raw, row_idx, row_errors, row_warnings)
        parsed_rows.append({
            'excel_row': row_idx, 'cleaned': cleaned, 'errors': row_errors,
            'warnings': row_warnings, 'raw_by_col': raw_by_col,
        })

    # 文件内重复检测
    seen_keys = {}
    for row in parsed_rows:
        if row['errors'] or not row['cleaned'].get('datetime'):
            continue
        key = _duplicate_key(config, row['cleaned'])
        if key in seen_keys:
            row['errors'].append(_row_error(
                row['excel_row'], '',
                f'与文件内第 {seen_keys[key]} 行重复（判断依据：{config["duplicate_label"]}），'
                '默认拒绝重复导入', ''))
        else:
            seen_keys[key] = row['excel_row']

    # 数据库重复预警（租户隔离）需要 request.user，
    # 由视图层调用 apply_db_duplicate_errors 后回填到 rows
    return {
        'business': business,
        'header_row': header_row,
        'header_labels': header_labels,
        'total_rows': len(parsed_rows),
        'rows': parsed_rows,
        'warnings': warnings,
    }


def apply_db_duplicate_errors(business, user, result):
    """将数据库重复检查结果回填到解析结果（validate 响应用）。"""
    duplicates = find_db_duplicate_rows(business, user, result['rows'])
    for row in result['rows']:
        message = duplicates.get(row['excel_row'])
        if message and not any(e['message'] == message for e in row['errors']):
            row['errors'].append(_row_error(row['excel_row'], '', message, ''))
    return result


def build_stats(result):
    """统计预校验结果：总行数/可导入数/错误数/警告数。"""
    error_count = sum(len(row['errors']) for row in result['rows'])
    warning_count = len(result.get('warnings') or []) + sum(
        len(row['warnings']) for row in result['rows'])
    valid_count = sum(1 for row in result['rows'] if not row['errors'])
    return {
        'total_rows': result['total_rows'],
        'valid_count': valid_count,
        'error_count': error_count,
        'warning_count': warning_count,
    }


# ====================================================================
# 错误报告
# ====================================================================

def build_error_report_workbook(business, result):
    """构建错误报告：与原文件列一致（含未识别列），追加 Excel行号 + 错误原因。"""
    header_labels = result['header_labels']
    col_indexes = sorted(header_labels.keys())
    headers = [header_labels[col_idx] for col_idx in col_indexes]
    headers += ['Excel行号', '错误原因']

    wb = Workbook()
    ws = wb.active
    ws.title = '导入错误报告'
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = Font(name='宋体', bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'

    out_row = 1
    for row in result['rows']:
        if not row['errors']:
            continue
        out_row += 1
        for col_idx, value in enumerate(
                (row['raw_by_col'].get(c, '') for c in col_indexes), start=1):
            _safe_cell(ws, out_row, col_idx, str(value))
        _safe_cell(ws, out_row, len(col_indexes) + 1, row['excel_row'])
        _safe_cell(ws, out_row, len(col_indexes) + 2,
                   '；'.join(error['message'] for error in row['errors']))

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

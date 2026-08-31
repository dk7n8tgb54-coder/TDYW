# Copyright: (c) OpenSpug Organization. https://github.com/openspug/spug
# Copyright (c) <spug.dev@gmail.com>
# Released under the AGPL-3.0 License.
"""
干扰管理导出服务

导出基于当前筛选条件下的全部数据，而非当前页。

附件支持：
- "附件"列：列出所有附件文件名（文本）
- "附件图片"列：嵌入第一张图片原图（仅图片类型，非缩略图）
- 有附件时导出 ZIP 压缩包：Excel（列表+原图嵌入）+ 附件/（所有原始文件含文档）
- 无附件时导出纯 Excel

ZIP 结构：
  干扰信息统计_xxx.zip
  ├── 干扰信息统计_xxx.xlsx
  └── 附件/
      ├── 001_频率123_汇报科室A/
      │   ├── 现场照片.jpg
      │   └── 干扰报告.pdf
      └── 002_频率456_汇报科室B/
          └── 照片.png
"""
import os
import io
import zipfile
import logging
from datetime import datetime
from collections import defaultdict
from urllib.parse import quote

from django.http import HttpResponse
from django.conf import settings
from django.views.generic import View

from libs import auth
from libs.export_utils import check_export_limit, build_export_error_response
from libs.tenant_utils import apply_tenant_filter
from apps.interference.models import (
    Interference, BridgeInterferenceRecord, AirInterferenceRecord,
)
from apps.evidence.models import EvidenceAttachment
from apps.logs.audit import record_audit_event

logger = logging.getLogger(__name__)

# 可嵌入 Excel 的图片扩展名
IMAGE_EXTENSIONS = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp')
# 嵌入图片数量上限（原图较大，限制数量避免 Excel 过大）
MAX_EMBED_IMAGES = 50
# 嵌入 Excel 的单张图片大小上限（超过则跳过嵌入，仅在 ZIP 中提供）
MAX_IMAGE_FILE_MB = 10
# 图片在 Excel 中的显示尺寸上限（像素），按原始比例缩放
MAX_DISPLAY_WIDTH = 480
MAX_DISPLAY_HEIGHT = 360

# 导出列定义：(字段, 表头)
# '__image__' 为特殊列：不写文本值，而是嵌入图片
EXCEL_COLUMNS = [
    ('export_serial', '序号'),
    ('frequency', '频率'),
    ('report_dept', '汇报科室'),
    ('datetime', '日期时间'),
    ('coordinates', '坐标'),
    ('interference_type', '干扰类型'),
    ('phenomenon', '现象'),
    ('flight_number', '航班号'),
    ('aircraft_type', '机型'),
    ('is_reported', '是否上报'),
    ('created_at', '创建时间'),
    ('attachment_names', '附件'),
    ('__image__', '附件图片'),
]

SHEET_NAME = '干扰信息统计'


def get_export_queryset(request):
    """按当前筛选条件查询数据，与前端 store 的过滤规则保持一致。"""
    qs = apply_tenant_filter(Interference.objects.filter(is_deleted=False), request.user)
    frequency = request.GET.get('frequency')
    if frequency:
        qs = qs.filter(frequency__icontains=frequency)
    report_dept = request.GET.get('report_dept')
    if report_dept:
        qs = qs.filter(report_dept__icontains=report_dept)
    interference_type = request.GET.get('interference_type')
    if interference_type:
        qs = qs.filter(interference_type__icontains=interference_type)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        qs = qs.filter(datetime__gte=start_date)
    if end_date:
        qs = qs.filter(datetime__lte=end_date + ' 23:59:59')
    return qs


def _build_filename_base(request):
    """构建不含扩展名的文件名基础部分。"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        scope = '%s-%s' % (start_date, end_date)
    else:
        scope = 'all'
    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    return '干扰信息统计_%s_%s' % (scope, now)


def _embed_image(ws, cell_ref, file_path, row_idx):
    """在指定单元格嵌入图片原图（非缩略图），成功返回 True。

    - 嵌入原始文件数据（不重新编码）
    - 显示尺寸按原始比例缩放到 MAX_DISPLAY_WIDTH x MAX_DISPLAY_HEIGHT 以内
    - 超过 MAX_IMAGE_FILE_MB 的图片跳过（避免 Excel 过大）
    """
    try:
        # 超大图片跳过嵌入
        file_size = os.path.getsize(file_path)
        if file_size > MAX_IMAGE_FILE_MB * 1024 * 1024:
            logger.info(f'图片过大跳过嵌入: {file_path} ({file_size / 1024 / 1024:.1f}MB)')
            return False

        from PIL import Image as PILImage
        from openpyxl.drawing.image import Image as XlImage

        # 获取原始尺寸用于显示比例
        pil_img = PILImage.open(file_path)
        orig_w, orig_h = pil_img.size
        pil_img.close()

        # 读取原始文件数据（不重新编码，保留原图质量）
        with open(file_path, 'rb') as f:
            stream = io.BytesIO(f.read())

        xl_img = XlImage(stream)

        # 按原始比例计算显示尺寸
        ratio = min(MAX_DISPLAY_WIDTH / orig_w, MAX_DISPLAY_HEIGHT / orig_h, 1.0)
        xl_img.width = max(int(orig_w * ratio), 50)
        xl_img.height = max(int(orig_h * ratio), 38)

        ws.add_image(xl_img, cell_ref)
        # 行高 = 显示高度(像素) * 0.75 (px->point 转换) + 5 (padding)
        ws.row_dimensions[row_idx].height = xl_img.height * 0.75 + 5
        return True
    except Exception as e:
        logger.error(f'嵌入图片失败: {file_path} {e}')
        return False


def _sanitize_zip_name(name):
    """清理文件/文件夹名中的非法字符（Windows 限制）。"""
    for ch in '/\\:*?"<>|':
        name = name.replace(ch, '_')
    return name.strip().rstrip('.')


def _write_excel_headers(ws, columns):
    """写表头行。"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(name='宋体', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (_key, title) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    return border


def _write_excel_data_rows(ws, columns, rows, border):
    """写数据行（跳过 __image__ 列）。"""
    from openpyxl.styles import Font, Alignment

    body_font = Font(name='宋体', size=10)
    body_align = Alignment(vertical='center', wrap_text=True)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, (key, _title) in enumerate(columns, start=1):
            if key == '__image__':
                continue
            value = row_data.get(key, '') if isinstance(row_data, dict) else getattr(row_data, key, '')
            if value is None:
                value = ''
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = border


def _embed_excel_images(ws, image_data, image_col_idx):
    """嵌入图片原图，返回已嵌入数量。"""
    from openpyxl.utils import get_column_letter

    embedded_count = 0
    if image_col_idx is None:
        return embedded_count

    for row_idx in sorted(image_data.keys()):
        if embedded_count >= MAX_EMBED_IMAGES:
            break
        for att in image_data[row_idx]:
            ext = (att.file_ext or '').lower()
            if not ext.startswith('.'):
                ext = '.' + ext
            if ext not in IMAGE_EXTENSIONS:
                continue
            full_path = os.path.join(settings.MEDIA_ROOT, att.file_path)
            if not os.path.exists(full_path):
                logger.warning(f'附件文件不存在: {full_path}')
                continue
            cell_ref = '%s%d' % (get_column_letter(image_col_idx), row_idx)
            if _embed_image(ws, cell_ref, full_path, row_idx):
                embedded_count += 1
            break
    return embedded_count


def _set_excel_column_widths(ws, columns, rows):
    """自适应列宽。"""
    from openpyxl.utils import get_column_letter

    for col_idx, (key, title) in enumerate(columns, start=1):
        if key == '__image__':
            ws.column_dimensions[get_column_letter(col_idx)].width = 40
            continue
        max_len = len(str(title)) * 2
        for row_data in rows[:200]:
            val = row_data.get(key, '') if isinstance(row_data, dict) else getattr(row_data, key, '')
            if val is not None and val != '':
                max_len = max(max_len, min(len(str(val)) * 2, 60))
        width = min(max(max_len + 4, 10), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _build_interference_excel(columns, rows, image_data, sheet_name=None):
    """构建干扰记录 Excel（含原图嵌入），返回 BytesIO。"""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name or SHEET_NAME

    border = _write_excel_headers(ws, columns)

    image_col_idx = None
    for col_idx, (key, _title) in enumerate(columns, start=1):
        if key == '__image__':
            image_col_idx = col_idx

    _write_excel_data_rows(ws, columns, rows, border)
    _embed_excel_images(ws, image_data, image_col_idx)
    _set_excel_column_widths(ws, columns, rows)

    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _build_zip_response(zip_filename, excel_data, record_ids, rows, attachments_by_object,
                        folder_label_fn=None):
    """构建 ZIP 响应：Excel + 附件原始文件目录。

    folder_label_fn: 可选，row dict -> 附件目录标签（不含序号）。
        为空时沿用旧干扰记录的 频率_汇报科室_日期时间 命名，行为不变。
    """
    zip_buf = io.BytesIO()
    used_names = set()

    excel_filename = zip_filename.rsplit('.', 1)[0] + '.xlsx'

    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # 1. 写入 Excel
        zf.writestr(excel_filename, excel_data.getvalue())
        used_names.add(excel_filename)

        # 2. 写入附件原始文件
        for idx, rid in enumerate(record_ids, start=1):
            attachments = attachments_by_object.get(rid, [])
            if not attachments:
                continue

            record = rows[idx - 1]
            if folder_label_fn is None:
                freq = _sanitize_zip_name(str(record.get('frequency', '')))
                dept = _sanitize_zip_name(str(record.get('report_dept', '')))
                dt = _sanitize_zip_name(str(record.get('datetime', '')))
                folder = '附件/%03d_%s_%s_%s' % (idx, freq, dept, dt)
            else:
                folder = '附件/%03d_%s' % (idx, _sanitize_zip_name(str(folder_label_fn(record))))

            for att in attachments:
                full_path = os.path.join(settings.MEDIA_ROOT, att.file_path)
                if not os.path.exists(full_path):
                    logger.warning(f'附件文件不存在，跳过: {full_path}')
                    continue
                # 构造 ZIP 内路径，处理重名
                file_name = _sanitize_zip_name(att.file_name)
                arcname = '%s/%s' % (folder, file_name)
                if arcname in used_names:
                    base, ext = os.path.splitext(file_name)
                    counter = 1
                    while '%s/%s_%d%s' % (folder, base, counter, ext) in used_names:
                        counter += 1
                    arcname = '%s/%s_%d%s' % (folder, base, counter, ext)
                used_names.add(arcname)
                zf.write(full_path, arcname)

    zip_buf.seek(0)
    response = HttpResponse(zip_buf.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = "attachment; filename*=UTF-8''%s" % quote(zip_filename)
    return response


class InterferenceExportView(View):
    """干扰信息导出

    - 无附件：导出纯 Excel
    - 有附件：导出 ZIP（Excel 含原图嵌入 + 附件原始文件目录）
    """

    @auth('interference.interference.view')
    def get(self, request):
        qs = get_export_queryset(request)
        count, error_resp = check_export_limit(qs)
        if error_resp:
            return error_resp
        if count == 0:
            return build_export_error_response('当前筛选条件下没有可导出的数据')

        records = qs.select_related('created_by', 'updated_by')

        # 第一遍遍历：构建行数据，收集 record_id
        rows = []
        record_ids = []
        for idx, obj in enumerate(records.iterator(), start=1):
            row = obj.to_dict()
            row['export_serial'] = idx
            row['attachment_names'] = ''
            rows.append(row)
            record_ids.append(str(obj.id))

        # 批量查询附件（避免 N+1）
        all_attachments = apply_tenant_filter(
            EvidenceAttachment.objects.filter(
                module='interference',
                object_type='interference',
                object_id__in=record_ids,
                is_deleted=False,
            ),
            request.user
        ).order_by('-uploaded_at')

        attachments_by_object = defaultdict(list)
        for att in all_attachments:
            attachments_by_object[att.object_id].append(att)

        # 填充附件信息到行数据
        image_data = {}
        total_attachments = 0
        for idx, rid in enumerate(record_ids, start=1):
            attachments = attachments_by_object.get(rid, [])
            names = [att.file_name for att in attachments]
            rows[idx - 1]['attachment_names'] = ', '.join(names) if names else ''
            if attachments:
                image_data[idx + 1] = attachments
                total_attachments += len(attachments)

        # 构建 Excel
        excel_buf = _build_interference_excel(EXCEL_COLUMNS, rows, image_data)

        filename_base = _build_filename_base(request)

        if total_attachments > 0:
            # 有附件：导出 ZIP
            zip_filename = filename_base + '.zip'
            record_audit_event(request, 'export', 'interference',
                               target_name=zip_filename,
                               detail={
                                   'count': len(rows),
                                   'format': 'zip',
                                   'attachments': total_attachments,
                                   'embedded_images': min(len(image_data), MAX_EMBED_IMAGES),
                               })
            return _build_zip_response(zip_filename, excel_buf, record_ids, rows, attachments_by_object)
        else:
            # 无附件：导出纯 Excel
            xlsx_filename = filename_base + '.xlsx'
            record_audit_event(request, 'export', 'interference',
                               target_name=xlsx_filename,
                               detail={'count': len(rows), 'format': 'xlsx'})
            response = HttpResponse(
                excel_buf.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = "attachment; filename*=UTF-8''%s" % quote(xlsx_filename)
            return response


# ==================== 双业务类型导出（地面/空中） ====================
#
# 列定义与各自表单字段及顺序一致；附件导出行为（文本列 + 原图嵌入 + ZIP 原始文件）
# 沿用旧干扰导出的安全机制，通过 folder_label_fn 差异化 ZIP 内目录命名。

BRIDGE_EXCEL_COLUMNS = [
    ('export_serial', '序号'),
    ('datetime', '日期时间'),
    ('flight_number', '航班号'),
    ('aircraft_no', '机号'),
    ('aircraft_type', '机型'),
    ('location', '位置/机位'),
    ('frequency', '频率'),
    ('phenomenon', '现象'),
    ('remark', '备注'),
    ('attachment_names', '附件'),
    ('__image__', '附件图片'),
]

AIR_EXCEL_COLUMNS = [
    ('export_serial', '序号'),
    ('datetime', '日期时间'),
    ('flight_number', '航班号'),
    ('aircraft_type', '机型'),
    ('route', '航线'),
    ('runway', '使用跑道'),
    ('approach_procedure', '使用进近程序'),
    ('alert_form', '被扰频率'),
    ('alert_altitude_text', '告警高度'),
    ('alert_segment', '告警航段'),
    ('duration_text', '持续时间'),
    ('phenomenon', '现象'),
    ('handling_method', '处置方式'),
    ('cause_analysis', '原因分析'),
    ('attachment_names', '附件'),
    ('__image__', '附件图片'),
]

BRIDGE_SHEET_NAME = '地面干扰记录'
AIR_SHEET_NAME = '空中干扰记录'


def _get_business_export_queryset(request, model, flight_field='flight_number'):
    """按当前筛选条件查询业务记录，与前端列表过滤规则保持一致。"""
    qs = apply_tenant_filter(model.objects.filter(is_deleted=False), request.user)
    flight_number = request.GET.get('flight_number')
    if flight_number:
        qs = qs.filter(**{f'{flight_field}__icontains': flight_number})
    status = request.GET.get('status')
    if status:
        qs = qs.filter(status=status)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        qs = qs.filter(datetime__gte=start_date)
    if end_date:
        qs = qs.filter(datetime__lte=end_date + ' 23:59:59')
    return qs


def _build_business_filename_base(request, prefix):
    """构建不含扩展名的文件名基础部分。"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        scope = '%s-%s' % (start_date, end_date)
    else:
        scope = 'all'
    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    return '%s_%s_%s' % (prefix, scope, now)


def _export_business_records(request, model, object_type, columns, sheet_name,
                             filename_prefix, folder_label_fn):
    """双业务记录导出公共实现。

    - 无附件：导出纯 Excel；
    - 有附件：导出 ZIP（Excel 含原图嵌入 + 附件原始文件目录）；
    - 附件查询含租户过滤，仅导出当前租户可见附件。
    """
    qs = _get_business_export_queryset(request, model)
    count, error_resp = check_export_limit(qs)
    if error_resp:
        return error_resp
    if count == 0:
        return build_export_error_response('当前筛选条件下没有可导出的数据')

    rows = []
    record_ids = []
    for idx, obj in enumerate(qs.iterator(), start=1):
        row = obj.to_view()
        # 日期时间业务精度到分钟（历史秒级数据展示为 :00 由分钟截断统一）
        if isinstance(row.get('datetime'), datetime):
            row['datetime'] = row['datetime'].strftime('%Y-%m-%d %H:%M')
        row['export_serial'] = idx
        row['attachment_names'] = ''
        rows.append(row)
        record_ids.append(str(obj.id))

    # 批量查询附件（避免 N+1）
    all_attachments = apply_tenant_filter(
        EvidenceAttachment.objects.filter(
            module='interference',
            object_type=object_type,
            object_id__in=record_ids,
            is_deleted=False,
        ),
        request.user,
    ).order_by('-uploaded_at')

    attachments_by_object = defaultdict(list)
    for att in all_attachments:
        attachments_by_object[att.object_id].append(att)

    image_data = {}
    total_attachments = 0
    for idx, rid in enumerate(record_ids, start=1):
        attachments = attachments_by_object.get(rid, [])
        names = [att.file_name for att in attachments]
        rows[idx - 1]['attachment_names'] = ', '.join(names) if names else ''
        if attachments:
            image_data[idx + 1] = attachments
            total_attachments += len(attachments)

    excel_buf = _build_interference_excel(columns, rows, image_data, sheet_name=sheet_name)
    filename_base = _build_business_filename_base(request, filename_prefix)

    if total_attachments > 0:
        zip_filename = filename_base + '.zip'
        record_audit_event(request, 'export', 'interference',
                           target_name=zip_filename,
                           detail={
                               'record_type': object_type,
                               'count': len(rows),
                               'format': 'zip',
                               'attachments': total_attachments,
                               'embedded_images': min(len(image_data), MAX_EMBED_IMAGES),
                           })
        return _build_zip_response(
            zip_filename, excel_buf, record_ids, rows, attachments_by_object,
            folder_label_fn=folder_label_fn)

    xlsx_filename = filename_base + '.xlsx'
    record_audit_event(request, 'export', 'interference',
                       target_name=xlsx_filename,
                       detail={'record_type': object_type, 'count': len(rows), 'format': 'xlsx'})
    response = HttpResponse(
        excel_buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename*=UTF-8''%s" % quote(xlsx_filename)
    return response


class InterferenceBridgeExportView(View):
    """地面无线电通信异常/干扰记录导出。"""

    @auth('interference.interference.view')
    def get(self, request):
        def folder_label_fn(row):
            return '%s_%s_%s' % (
                row.get('flight_number', ''), row.get('location', ''), row.get('datetime', ''))

        return _export_business_records(
            request, BridgeInterferenceRecord, 'bridge_interference',
            BRIDGE_EXCEL_COLUMNS, BRIDGE_SHEET_NAME, '地面干扰信息', folder_label_fn)


class InterferenceAirExportView(View):
    """空中干扰记录导出。"""

    @auth('interference.interference.view')
    def get(self, request):
        def folder_label_fn(row):
            return '%s_%s_%s' % (
                row.get('flight_number', ''), row.get('route', ''), row.get('datetime', ''))

        return _export_business_records(
            request, AirInterferenceRecord, 'air_interference',
            AIR_EXCEL_COLUMNS, AIR_SHEET_NAME, '空中干扰信息', folder_label_fn)

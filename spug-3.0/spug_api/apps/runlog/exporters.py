# Copyright: (c) OpenSpug Organization. https://github.com/openspug/spug
# Copyright: (c) <spug.dev@gmail.com>
# Released under the AGPL-3.0 License.
"""
运行日志 Excel 导出服务

采用"事件行 + 动态行"主从结构，同一 Sheet 内交替输出：
- 事件行：填写事件概要字段，动态相关字段留空，加粗 + 浅色背景
- 动态行：填写动态字段，事件概要字段（除事件序号）留空，普通字体

不改动现有 PDF 导出逻辑，Excel 与 PDF 并存。
导出基于当前筛选条件下的全部数据，而非当前页。
"""
import io
import logging
from datetime import datetime, timedelta
from libs.date_utils import date_range_filter
from urllib.parse import quote

from django.http import HttpResponse, JsonResponse
from django.views.generic import View
from libs import auth
from libs.export_utils import check_export_limit, build_export_error_response
from libs.tenant_utils import apply_tenant_filter
from apps.runlog.models import RunLog, RunLogUpdate

logger = logging.getLogger(__name__)

SHEET_NAME = '运行日志明细'

# 统一表头：(字段key, 表头)
# 注：附件能力已下线，导出不再包含附件相关字段
COLUMNS = [
    ('row_type', '行类型'),
    ('event_index', '事件序号'),
    ('event_title', '事件标题'),
    ('event_type', '事件类型'),
    ('system_name', '系统名称'),
    ('severity_text', '事件级别'),
    ('status_text', '当前状态'),
    ('responsible_user_name', '责任人'),
    ('created_at', '创建时间'),
    ('updated_at', '更新时间'),
    ('update_count', '动态条数'),
    ('update_index', '动态序号'),
    ('update_date', '动态日期'),
    ('recorder', '记录人'),
    ('duty_person', '值班人'),
    ('detail_content', '动态内容'),
]

EVENT_ROW_TYPE = '事件'
UPDATE_ROW_TYPE = '动态'


def _apply_filters(qs, request):
    """复用列表与 PDF 接口的筛选条件，保证导出结果与列表筛选一致。"""
    status = request.GET.get('status')
    if status:
        qs = qs.filter(status=status)
    severity = request.GET.get('severity')
    if severity:
        qs = qs.filter(severity=severity)
    event_type = request.GET.get('event_type')
    if event_type:
        qs = qs.filter(event_type=event_type)
    responsible_user_name = request.GET.get('responsible_user_name')
    if responsible_user_name:
        qs = qs.filter(responsible_user_name__icontains=responsible_user_name)
    system_name = request.GET.get('system_name')
    if system_name:
        qs = qs.filter(system_name__icontains=system_name)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    qs = date_range_filter(qs, 'created_at', start_date, end_date)
    return qs


def _build_rows(events):
    """构建主从结构行数据：事件行 + 该事件的动态行"""
    rows = []
    for event_idx, event in enumerate(events, start=1):
        event_view = event.to_view()
        updates = getattr(event, '_export_updates', None) or []
        update_total = len(updates)

        # 事件行
        rows.append({
            'row_type': EVENT_ROW_TYPE,
            'event_index': event_idx,
            'event_title': event_view.get('event_title', ''),
            'event_type': event_view.get('event_type', ''),
            'system_name': event_view.get('system_name', ''),
            'severity_text': event_view.get('severity_text', ''),
            'status_text': event_view.get('status_text', ''),
            'responsible_user_name': event_view.get('responsible_user_name', '') or '',
            'created_at': event_view.get('created_at', ''),
            'updated_at': event_view.get('updated_at', '') or '',
            'update_count': update_total,
        })

        # 动态行
        if not updates:
            rows.append({
                'row_type': UPDATE_ROW_TYPE,
                'event_index': event_idx,
                'detail_content': '暂无动态记录',
            })
        else:
            for upd_idx, upd in enumerate(updates, start=1):
                upd_view = upd.to_view()
                rows.append({
                    'row_type': UPDATE_ROW_TYPE,
                    'event_index': event_idx,
                    'update_index': upd_idx,
                    'update_date': upd_view.get('update_date', ''),
                    'recorder': upd_view.get('recorder', ''),
                    'duty_person': upd_view.get('duty_person', '') or '',
                    'detail_content': upd_view.get('detail_content', ''),
                })
    return rows


def _build_excel_response(filename, rows):
    """
    构建 Excel 响应，主从结构差异化样式：
    - 表头：白字 + 深蓝背景 + 居中
    - 事件行：加粗 + 浅蓝背景
    - 动态行：普通字体，动态内容列缩进显示
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # 样式
    header_font = Font(name='宋体', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    event_font = Font(name='宋体', bold=True, size=10)
    event_fill = PatternFill(start_color='E6F7FF', end_color='E6F7FF', fill_type='solid')

    update_font = Font(name='宋体', size=10)
    update_align = Alignment(vertical='center', wrap_text=True)

    thin = Side(border_style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 写表头
    for col_idx, (_key, title) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # 写数据行
    for row_idx, row_data in enumerate(rows, start=2):
        is_event = row_data.get('row_type') == EVENT_ROW_TYPE
        for col_idx, (key, _title) in enumerate(COLUMNS, start=1):
            value = row_data.get(key, '')
            if value is None:
                value = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if is_event:
                cell.font = event_font
                cell.fill = event_fill
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            else:
                cell.font = update_font
                # 动态内容列缩进
                if key == 'detail_content':
                    cell.alignment = Alignment(vertical='center', wrap_text=True, indent=1)
                else:
                    cell.alignment = update_align

    # 列宽
    col_widths = {
        'row_type': 8, 'event_index': 8, 'event_title': 24, 'event_type': 12,
        'system_name': 14, 'severity_text': 8, 'status_text': 10,
        'responsible_user_name': 10, 'created_at': 18, 'updated_at': 18,
        'update_count': 8, 'update_index': 8, 'update_date': 12,
        'recorder': 10, 'duty_person': 10, 'detail_content': 50,
    }
    for col_idx, (key, _title) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(key, 12)

    # 冻结表头
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename*=UTF-8''%s" % quote(filename)
    return response


def _build_filename():
    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    return '运行日志明细_%s.xlsx' % now


class RunLogExcelExportView(View):
    """运行日志 Excel 明细导出（主从结构：事件行 + 动态行）"""

    @auth('runlog.runlog.view')
    def get(self, request):
        qs = apply_tenant_filter(RunLog.objects.all(), request.user)
        qs = _apply_filters(qs, request)
        qs = qs.order_by('-created_at', '-id')

        count, error_resp = check_export_limit(qs)
        if error_resp:
            return error_resp
        if count == 0:
            return build_export_error_response('当前筛选条件下没有可导出的数据')

        # 预加载事件 + 关联动态（避免 N+1）
        events = list(qs.select_related('created_by', 'updated_by'))
        event_ids = [e.id for e in events]
        all_updates = list(
            RunLogUpdate.objects.filter(runlog_id__in=event_ids)
            .order_by('runlog_id', 'update_date', 'sequence', 'id')
            .select_related('created_by')
        )
        updates_map = {}
        for upd in all_updates:
            updates_map.setdefault(upd.runlog_id, []).append(upd)
        for event in events:
            event._export_updates = updates_map.get(event.id, [])

        rows = _build_rows(events)
        filename = _build_filename()
        return _build_excel_response(filename, rows)

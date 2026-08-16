# Copyright: (c) OpenSpug Organization. https://github.com/openspug
# Copyright: (c) <spug.dev@gmail.com>
# Released under the AGPL-3.0 License.
"""系统升级模块缺陷复现测试（defect_reproduction）

候选缺陷清单（审查 upgrade 前后端代码得出，断言一律指向期望的正确行为，
修复前失败即证明缺陷真实存在，修复后必须全部通过）：

B1  更新接口对非法计划升级时间未校验未捕获：
    RecordService.update_record 无 try/except，upgrade_time 传入
    'not-a-date' / '' / '2026-02-30 10:00:00' 等非法值时 Django
    DateTimeField 抛 ValidationError → HTTP 500。创建接口虽有 try/except
    不会 500，但返回英文技术性错误信息。
    期望：创建与更新均返回 json_response(error='计划升级时间格式无效...')，不落库。

B2  更新接口校验缺失（与创建不对称）：
    validate_update 只校验状态流转，PUT 可把 title/system/owner/
    upgrade_content 清空为 ''，可把 upgrade_type 写成任意非法枚举值。
    期望：与创建一致，拒绝空必填字段和非法枚举值。

B3  已完成记录出现待执行步骤后主表状态不回退：
    _check_and_update_record_status 规定"存在待执行步骤且状态为已完成
    → 回退处理中"，但 add_manual_step 与 PlanService.apply_to_record
    均不调用该检查，新增/替换 pending 步骤后记录仍显示"已完成"。
    期望：出现待执行步骤后主表状态回退为"处理中"。

B4  skip 动作已声明但未实现：
    RecordStepUpdateView 帮助文本声明支持 complete/skip/reset，
    模型定义 STEP_STATUS_SKIPPED、步骤统计/阶段完成/自动完成逻辑均
    已支持 skipped，但 update_step_status 不接受 'skip' 动作。
    期望：action='skip' 将步骤置为已跳过，并参与阶段完成与自动完成判定。

B5  Excel 导出包含模型已不存在的 version 字段：
    UpgradeRecord 无 version 字段，导出列 ('version', '版本') 对每行
    输出空字符串，产生误导性空列。
    期望：导出表头不再包含"版本"空列。

B6  手动添加步骤序号用 count() 计算：
    add_manual_step 以未删除步骤数 +1 作为新序号，存在软删除步骤时
    会与现存步骤序号重复。
    期望：新序号取当前最大 sequence + 1。

B7  回退目标阶段不存在时静默重置全部步骤：
    add_log 对 rollback 只校验 target_action 非空，传入不存在的阶段时
    _mark_rollback_phases_failed/_reset_steps_for_rollback 回退到 idx=0，
    把所有阶段全部标记失败并重置步骤。
    期望：回退目标必须是该记录现存步骤的阶段，否则返回业务错误且不产生副作用。
"""
import io
import json
from datetime import datetime

from django.test import TestCase

from apps.upgrade.models import UpgradeRecord
from apps.upgrade.models_checklist import UpgradeRecordStep
from apps.upgrade.models_status_log import UpgradeStatusLog
from apps.upgrade.models_template import UpgradeTemplate, UpgradePlanStep
from apps.utils.test_helpers import make_user, make_client, setup_test_env

PERMS = [
    'upgrade.upgrade.view', 'upgrade.upgrade.add', 'upgrade.upgrade.edit',
    'upgrade.upgrade.del', 'upgrade.upgrade.step_reset',
    'upgrade.upgrade.step_del', 'upgrade.statistics.view',
]


class UpgradeDefectBase(TestCase):
    """缺陷复现测试公共基类"""

    def setUp(self):
        setup_test_env(self)
        self.user = make_user('upg_defect', PERMS)
        self.client = make_client(self.user)
        self.record = UpgradeRecord.objects.create(
            tenant_id=self.user.tenant_id, title='缺陷验证单',
            system='测试系统', upgrade_type='功能升级', status='处理中',
            created_by=self.user)

    def add_step(self, title, phase='', sequence=None):
        return UpgradeRecordStep.objects.create(
            tenant_id=self.user.tenant_id, upgrade_id=self.record.id,
            checklist_id=0, phase=phase, title=title,
            sequence=sequence if sequence is not None else 1)

    def complete_step(self, step):
        resp = self.put_json(f'/upgrade/record-steps/{step.id}/update/',
                             {'action': 'complete'})
        data = resp.json()
        self.assertFalse(data.get('error'), data)

    def put_json(self, url, payload):
        return self.client.put(url, data=json.dumps(payload),
                               content_type='application/json')

    def post_json(self, url, payload):
        return self.client.post(url, data=json.dumps(payload),
                                content_type='application/json')


# ============ B1：非法计划升级时间 ============

class InvalidUpgradeTimeTests(UpgradeDefectBase):
    """B1：非法计划升级时间应返回业务错误而非 500/技术性错误"""

    def test_update_invalid_time_string_rejected(self):
        resp = self.put_json(f'/upgrade/records/{self.record.id}/update/',
                             {'upgrade_time': 'not-a-date'})
        data = resp.json()
        self.assertIn('计划升级时间', data.get('error', ''))
        self.record.refresh_from_db()
        self.assertIsNone(self.record.upgrade_time)

    def test_update_empty_time_rejected(self):
        resp = self.put_json(f'/upgrade/records/{self.record.id}/update/',
                             {'upgrade_time': ''})
        data = resp.json()
        self.assertIn('计划升级时间', data.get('error', ''))
        self.record.refresh_from_db()
        self.assertIsNone(self.record.upgrade_time)

    def test_update_impossible_date_rejected(self):
        resp = self.put_json(f'/upgrade/records/{self.record.id}/update/',
                             {'upgrade_time': '2026-02-30 10:00:00'})
        data = resp.json()
        self.assertIn('计划升级时间', data.get('error', ''))
        self.record.refresh_from_db()
        self.assertIsNone(self.record.upgrade_time)

    def test_update_valid_time_accepted(self):
        resp = self.put_json(f'/upgrade/records/{self.record.id}/update/',
                             {'upgrade_time': '2026-03-01 08:00:00'})
        data = resp.json()
        self.assertFalse(data.get('error'), data)
        self.record.refresh_from_db()
        self.assertEqual(self.record.upgrade_time, datetime(2026, 3, 1, 8, 0, 0))

    def test_create_invalid_time_rejected(self):
        payload = {
            'title': 'B1-创建非法时间', 'system': '测试系统',
            'upgrade_type': '功能升级', 'upgrade_time': 'not-a-date',
            'owner': '负责人', 'upgrade_content': '内容',
        }
        resp = self.post_json('/upgrade/records/create/', payload)
        data = resp.json()
        self.assertIn('计划升级时间', data.get('error', ''))
        self.assertFalse(UpgradeRecord.objects.filter(title='B1-创建非法时间').exists())


# ============ B2：更新接口校验缺失 ============

class UpdateValidationTests(UpgradeDefectBase):
    """B2：更新应与创建一致地拒绝空必填字段和非法枚举值"""

    def test_update_blank_title_rejected(self):
        resp = self.put_json(f'/upgrade/records/{self.record.id}/update/',
                             {'title': ''})
        self.assertIn('请填写', resp.json().get('error', ''))
        self.record.refresh_from_db()
        self.assertEqual(self.record.title, '缺陷验证单')

    def test_update_blank_required_fields_rejected(self):
        for field in ('system', 'owner', 'upgrade_content'):
            resp = self.put_json(f'/upgrade/records/{self.record.id}/update/',
                                 {field: ''})
            self.assertTrue(resp.json().get('error'), field)
        self.record.refresh_from_db()
        self.assertEqual(self.record.system, '测试系统')

    def test_update_invalid_upgrade_type_rejected(self):
        resp = self.put_json(f'/upgrade/records/{self.record.id}/update/',
                             {'upgrade_type': '乱写类型'})
        self.assertIn('升级类型无效', resp.json().get('error', ''))
        self.record.refresh_from_db()
        self.assertEqual(self.record.upgrade_type, '功能升级')

    def test_update_invalid_status_rejected(self):
        resp = self.put_json(f'/upgrade/records/{self.record.id}/update/',
                             {'status': '乱写状态'})
        self.assertTrue(resp.json().get('error'))
        self.record.refresh_from_db()
        self.assertEqual(self.record.status, '处理中')

    def test_update_valid_owner_change_accepted(self):
        resp = self.put_json(f'/upgrade/records/{self.record.id}/update/',
                             {'owner': '新负责人'})
        data = resp.json()
        self.assertFalse(data.get('error'), data)
        self.record.refresh_from_db()
        self.assertEqual(self.record.owner, '新负责人')


# ============ B3：已完成记录出现待执行步骤后状态不回退 ============

class CompletedRecordPendingStepTests(UpgradeDefectBase):
    """B3：已完成记录出现待执行步骤后主表状态应回退为处理中"""

    def _complete_all(self, steps):
        for s in steps:
            self.complete_step(s)
        self.record.refresh_from_db()
        # 前置：所有步骤完成后主表自动置为已完成
        self.assertEqual(self.record.status, '已完成')

    def test_add_pending_step_reverts_status(self):
        s1 = self.add_step('步骤一', phase='阶段A')
        s2 = self.add_step('步骤二', phase='阶段A')
        self._complete_all([s1, s2])

        resp = self.post_json(f'/upgrade/records/{self.record.id}/steps/add/',
                              {'title': '补增步骤'})
        data = resp.json()
        self.assertFalse(data.get('error'), data)

        self.record.refresh_from_db()
        self.assertEqual(self.record.status, '处理中')
        self.assertTrue(UpgradeRecordStep.objects.filter(
            upgrade_id=self.record.id, title='补增步骤', status='pending').exists())

    def test_apply_plan_replace_reverts_status(self):
        s1 = self.add_step('旧步骤', phase='阶段A')
        self._complete_all([s1])

        tpl = UpgradeTemplate.objects.create(
            tenant_id=self.user.tenant_id, name='B3缺陷方案', created_by=self.user)
        UpgradePlanStep.objects.create(
            tenant_id=self.user.tenant_id, template_id=tpl.id,
            phase='阶段A', title='新方案步骤', sequence=1)

        resp = self.post_json(f'/upgrade/plans/{tpl.id}/apply/',
                              {'upgrade_id': self.record.id, 'replace': True})
        data = resp.json()
        self.assertFalse(data.get('error'), data)

        self.record.refresh_from_db()
        self.assertEqual(self.record.status, '处理中')


# ============ B4：skip 动作未实现 ============

class SkipActionTests(UpgradeDefectBase):
    """B4：action='skip' 应将步骤置为已跳过并参与完成度判定"""

    def test_skip_marks_step_and_counts_done(self):
        s = self.add_step('可跳过步骤', phase='阶段A')
        resp = self.put_json(f'/upgrade/record-steps/{s.id}/update/',
                             {'action': 'skip'})
        data = resp.json()
        self.assertFalse(data.get('error'), data)

        s.refresh_from_db()
        self.assertEqual(s.status, 'skipped')
        self.assertTrue(s.completed_by)

        resp = self.client.get(f'/upgrade/records/{self.record.id}/steps/')
        stats = resp.json()['data']['stats']
        self.assertEqual(stats['skipped'], 1)
        self.assertEqual(stats['pending'], 0)

        # 跳过计入完成度：无待执行步骤 → 主表自动完成 + 写 phase_done
        self.record.refresh_from_db()
        self.assertEqual(self.record.status, '已完成')
        self.assertTrue(UpgradeStatusLog.objects.filter(
            upgrade_id=self.record.id, action='phase_done',
            phase='阶段A', outcome='done').exists())

    def test_skip_with_remark_saved(self):
        s = self.add_step('带备注跳过', phase='')
        resp = self.put_json(f'/upgrade/record-steps/{s.id}/update/',
                             {'action': 'skip', 'remark': '该步骤不适用'})
        data = resp.json()
        self.assertFalse(data.get('error'), data)
        s.refresh_from_db()
        self.assertEqual(s.status, 'skipped')
        self.assertEqual(s.remark, '该步骤不适用')

    def test_reset_skipped_step_back_to_pending(self):
        s = self.add_step('跳过后重置', phase='阶段A')
        resp = self.put_json(f'/upgrade/record-steps/{s.id}/update/',
                             {'action': 'skip'})
        self.assertFalse(resp.json().get('error'))

        resp = self.put_json(f'/upgrade/record-steps/{s.id}/update/',
                             {'action': 'reset'})
        data = resp.json()
        self.assertFalse(data.get('error'), data)
        s.refresh_from_db()
        self.assertEqual(s.status, 'pending')


# ============ B5：导出版本空列 ============

class ExportColumnTests(UpgradeDefectBase):
    """B5：导出不应包含模型已不存在的 version 空列"""

    def test_export_has_no_dead_version_column(self):
        resp = self.client.get('/upgrade/records/export/')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertNotIn('版本', headers)
        self.assertIn('标题', headers)
        # 至少一行数据（本测试创建的记录）
        self.assertGreaterEqual(ws.max_row, 2)


# ============ B6：手动步骤序号重复 ============

class ManualStepSequenceTests(UpgradeDefectBase):
    """B6：手动添加步骤序号应取 max(sequence)+1，软删除后不重复"""

    def test_sequence_uses_max_after_soft_delete(self):
        s1 = self.add_step('步骤一', sequence=1)
        self.add_step('步骤二', sequence=2)
        self.add_step('步骤三', sequence=3)

        resp = self.client.delete(f'/upgrade/record-steps/{s1.id}/delete/')
        self.assertFalse(resp.json().get('error'))

        resp = self.post_json(f'/upgrade/records/{self.record.id}/steps/add/',
                              {'title': '新步骤'})
        data = resp.json()
        self.assertFalse(data.get('error'), data)

        new_step = UpgradeRecordStep.objects.get(title='新步骤')
        self.assertEqual(new_step.sequence, 4)


# ============ B7：回退目标阶段校验 ============

class RollbackTargetValidationTests(UpgradeDefectBase):
    """B7：回退目标阶段不存在时应拒绝且不产生副作用"""

    def test_rollback_unknown_target_rejected(self):
        s = self.add_step('阶段A步骤', phase='阶段A')
        self.complete_step(s)
        self.record.refresh_from_db()
        # 前置：唯一步骤完成后主表自动置为已完成
        self.assertEqual(self.record.status, '已完成')

        resp = self.post_json(f'/upgrade/records/{self.record.id}/status-logs/',
                              {'action': 'rollback', 'target_action': '不存在的阶段'})
        self.assertTrue(resp.json().get('error'))

        # 无副作用：步骤不被重置、phase_done 不被改失败、主表不被联动
        s.refresh_from_db()
        self.assertEqual(s.status, 'completed')
        self.record.refresh_from_db()
        self.assertEqual(self.record.status, '已完成')
        log = UpgradeStatusLog.objects.get(
            upgrade_id=self.record.id, action='phase_done')
        self.assertEqual(log.outcome, 'done')
        self.assertFalse(UpgradeStatusLog.objects.filter(
            upgrade_id=self.record.id, action='rollback').exists())

    def test_rollback_valid_target_still_works(self):
        sa = self.add_step('阶段A步骤', phase='阶段A')
        sb = self.add_step('阶段B步骤', phase='阶段B')
        self.complete_step(sa)
        self.complete_step(sb)

        resp = self.post_json(f'/upgrade/records/{self.record.id}/status-logs/',
                              {'action': 'rollback', 'target_action': '阶段B'})
        data = resp.json()
        self.assertFalse(data.get('error'), data)

        sa.refresh_from_db()
        sb.refresh_from_db()
        self.assertEqual(sa.status, 'completed')
        self.assertEqual(sb.status, 'pending')
        self.record.refresh_from_db()
        self.assertEqual(self.record.status, '已回退')

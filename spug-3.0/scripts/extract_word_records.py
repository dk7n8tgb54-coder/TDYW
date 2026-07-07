#!/usr/bin/env python
"""Extract operation/disposal records from structured Word documents.

The target documents use a repeated two-row pattern:

1. Metadata row: system, date, duty staff, recorder.
2. Detail row: "处置记录" plus the detailed operation text.

Word merged cells are exposed by python-docx as repeated cell text, so this
script deliberately de-duplicates repeated cells and reads values after labels.
It writes an Excel preview for human review. It does not write application data.
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterable

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LABELS = (
    "系统",
    "日期",
    "当日处置值班人员",
    "值班人员",
    "记录人",
    "处置记录",
    "处置",
    "记录",
)

LABEL_ONLY_VALUES = set(LABELS)


@dataclass
class ExtractedRecord:
    source_file: str
    table_index: int
    meta_row: int
    detail_row: int
    system_name: str
    record_date: str
    duty_person: str
    recorder: str
    event_type: str
    event_title: str
    detail_content: str
    fault_phenomenon: str
    handling_process: str
    handling_result: str
    confidence: str
    notes: str


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", " ", value).strip()


def compact(value: str | None) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def dedupe_consecutive(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    last = None
    for value in values:
        text = normalize_text(value)
        if text and text != last:
            result.append(text)
            last = text
    return result


def looks_like_label(value: str) -> bool:
    text = compact(value)
    if not text:
        return True
    return text in LABEL_ONLY_VALUES


def value_after_label(cells: list[str], label_options: tuple[str, ...]) -> str:
    label_compacts = tuple(compact(x) for x in label_options)
    all_label_compacts = tuple(compact(x) for x in LABELS)

    for index, cell in enumerate(cells):
        cell_compact = compact(cell)
        if not any(label in cell_compact for label in label_compacts):
            continue
        for candidate in cells[index + 1 :]:
            candidate_text = normalize_text(candidate)
            candidate_compact = compact(candidate_text)
            if not candidate_text:
                continue
            if candidate_compact in all_label_compacts:
                continue
            return candidate_text
    return ""


def parse_date(value: str) -> str:
    text = normalize_text(value)
    patterns = (
        r"(?P<y>20\d{2})[.\-/年](?P<m>\d{1,2})[.\-/月](?P<d>\d{1,2})",
        r"(?P<y>20\d{2})(?P<m>\d{2})(?P<d>\d{2})",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            year = int(match.group("y"))
            month = int(match.group("m"))
            day = int(match.group("d"))
            return f"{year:04d}-{month:02d}-{day:02d}"
    return text


def extract_detail_content(cells: list[str]) -> str:
    candidates = [x for x in dedupe_consecutive(cells) if not looks_like_label(x)]
    if not candidates:
        return ""
    return max(candidates, key=len)


def split_detail(content: str) -> tuple[str, str, str]:
    markers = [
        ("fault", "故障现象"),
        ("process", "处置流程"),
        ("result", "处置结果"),
    ]
    matches = []
    for key, label in markers:
        match = re.search(rf"{label}\s*[:：]", content)
        if match:
            matches.append((match.start(), match.end(), key))
    if not matches:
        return "", "", ""

    matches.sort()
    parts = {"fault": "", "process": "", "result": ""}
    for index, (start, end, key) in enumerate(matches):
        next_start = matches[index + 1][0] if index + 1 < len(matches) else len(content)
        parts[key] = content[end:next_start].strip()
    return parts["fault"], parts["process"], parts["result"]


def guess_event_type(system_name: str, content: str) -> str:
    text = f"{system_name} {content}"
    if "干扰" in text:
        return "无线电干扰"
    if "升级" in text or "更新" in text:
        return "系统升级"
    if "故障" in text or "告警" in text or "报警" in text:
        return "设备故障"
    if "甚高频" in text or "频率" in text:
        return "甚高频处置"
    return "运行处置"


def build_title(system_name: str, content: str) -> str:
    cleaned = re.sub(r"(故障现象|处置流程|处置结果)\s*[:：]", "", content)
    first_sentence = re.split(r"[。；;]", cleaned, maxsplit=1)[0].strip()
    if not first_sentence:
        first_sentence = cleaned.strip()
    title = f"{system_name} - {first_sentence}" if system_name else first_sentence
    return title[:100]


def extract_records_from_docx(path: Path) -> list[ExtractedRecord]:
    document = Document(path)
    records: list[ExtractedRecord] = []

    for table_index, table in enumerate(document.tables, start=1):
        pending_meta: dict[str, str | int] | None = None
        for row_index, row in enumerate(table.rows, start=1):
            cells = [normalize_text(cell.text) for cell in row.cells]
            row_compact = compact(" ".join(cells))

            if "系统" in row_compact and "日期" in row_compact:
                system_name = value_after_label(cells, ("系统",))
                record_date = parse_date(value_after_label(cells, ("日期",)))
                if not system_name and not record_date:
                    pending_meta = None
                    continue
                pending_meta = {
                    "source_file": path.name,
                    "table_index": table_index,
                    "meta_row": row_index,
                    "system_name": system_name,
                    "record_date": record_date,
                    "duty_person": value_after_label(cells, ("当日处置值班人员", "值班人员")),
                    "recorder": value_after_label(cells, ("记录人",)),
                }
                continue

            if pending_meta is None:
                continue

            has_detail_label = "处置记录" in row_compact or (
                "处置" in row_compact and "记录" in row_compact
            )
            content = extract_detail_content(cells)
            if not content:
                pending_meta = None
                continue
            if not has_detail_label and len(content) < 10:
                continue

            fault, process, result = split_detail(content)
            notes = []
            for required in ("system_name", "record_date"):
                if not pending_meta.get(required):
                    notes.append(f"缺少{required}")
            if not content:
                notes.append("缺少处置记录")

            system_name = str(pending_meta.get("system_name") or "")
            event_type = guess_event_type(system_name, content)
            confidence = "高" if not notes else "中"
            records.append(
                ExtractedRecord(
                    source_file=str(pending_meta["source_file"]),
                    table_index=int(pending_meta["table_index"]),
                    meta_row=int(pending_meta["meta_row"]),
                    detail_row=row_index,
                    system_name=system_name,
                    record_date=str(pending_meta.get("record_date") or ""),
                    duty_person=str(pending_meta.get("duty_person") or ""),
                    recorder=str(pending_meta.get("recorder") or ""),
                    event_type=event_type,
                    event_title=build_title(system_name, content),
                    detail_content=content,
                    fault_phenomenon=fault,
                    handling_process=process,
                    handling_result=result,
                    confidence=confidence,
                    notes="; ".join(notes),
                )
            )
            pending_meta = None

    return records


def iter_docx_inputs(input_path: Path) -> list[Path]:
    if input_path.is_file():
        return [input_path]
    if input_path.is_dir():
        return sorted(
            x for x in input_path.rglob("*.docx") if not x.name.startswith("~$")
        )
    raise FileNotFoundError(f"Input path does not exist: {input_path}")


def write_preview_workbook(records: list[ExtractedRecord], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入预览"

    headers = [
        "来源文件",
        "表格序号",
        "元信息行",
        "详情行",
        "系统/设备",
        "日期",
        "值班人员",
        "记录人",
        "事件类型建议",
        "事件标题建议",
        "处置记录全文",
        "故障现象",
        "处置流程",
        "处置结果",
        "置信度",
        "备注",
    ]
    sheet.append(headers)
    for record in records:
        data = asdict(record)
        sheet.append(
            [
                data["source_file"],
                data["table_index"],
                data["meta_row"],
                data["detail_row"],
                data["system_name"],
                data["record_date"],
                data["duty_person"],
                data["recorder"],
                data["event_type"],
                data["event_title"],
                data["detail_content"],
                data["fault_phenomenon"],
                data["handling_process"],
                data["handling_result"],
                data["confidence"],
                data["notes"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin_gray)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    widths = {
        "A": 32,
        "B": 10,
        "C": 10,
        "D": 10,
        "E": 18,
        "F": 12,
        "G": 24,
        "H": 12,
        "I": 14,
        "J": 42,
        "K": 72,
        "L": 42,
        "M": 42,
        "N": 42,
        "O": 10,
        "P": 24,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[5].number_format = "yyyy-mm-dd"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    info = workbook.create_sheet("解析说明")
    info_rows = [
        ["项目", "说明"],
        ["用途", "从历史 Word 工作记录中提取结构化预览，供人工确认后再导入系统。"],
        ["安全性", "本脚本只读取 Word 并生成 Excel，不写数据库。"],
        ["解析规则", "识别包含“系统/日期/值班人员/记录人”的元信息行，并与下一条“处置记录”行配对。"],
        ["合并单元格", "Word 合并单元格会重复显示，本脚本会自动跳过连续重复值。"],
        ["建议导入模块", "多数记录可映射到运行日志；含故障、干扰、升级等关键词时会给出事件类型建议。"],
        ["记录数", len(records)],
    ]
    for row in info_rows:
        info.append(row)
    info.column_dimensions["A"].width = 18
    info.column_dimensions["B"].width = 90
    for cell in info[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in info.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract structured operation records from .docx files."
    )
    parser.add_argument("input", help="Input .docx file or directory")
    parser.add_argument(
        "-o",
        "--output",
        default="outputs/word_import/word_records_preview.xlsx",
        help="Output .xlsx preview path",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    records: list[ExtractedRecord] = []
    for docx_path in iter_docx_inputs(input_path):
        records.extend(extract_records_from_docx(docx_path))

    write_preview_workbook(records, output_path)
    print(f"Extracted {len(records)} records")
    print(f"Preview workbook: {output_path.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
